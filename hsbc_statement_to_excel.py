#!/usr/bin/env python3
"""Convert scanned HSBC One statement PDFs into Excel rows using OCR.

This script preserves your existing workbook format by writing transactions into a
sheet with the same column layout:
A T-Date | B V-Date | C Description | D CQ# | E CNY | F GBP | G JPY S/A | H HKD S/A | I USD S/A
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Font
from hsbc_general_parser import parse_statement_text as parse_general_statement_text

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover - optional fallback parser dependency
    PdfReader = None

DATE_RE = re.compile(r"^(\d{1,2})\s*([A-Za-z]{3})$")
FULL_DATE_RE = re.compile(r"^(\d{1,2})\s+([A-Za-z]{3})\s+(\d{2,4})$")
DATE_WITH_CCY_RE = re.compile(r"^(USD|GBP|JPY|CNY)\s*(\d{1,2}\s*[A-Za-z]{3})?$")
NUMBER_RE = re.compile(r"-?\d[\d,]*\.?\d*")
STATEMENT_DATE_RE = re.compile(r"(\d{1,2})\s+([A-Za-z]{3,9})\s+(20\d{2})")

MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

CURRENCY_COLUMN = {
    "CNY": 5,
    "GBP": 6,
    "JPY": 7,
    "HKD": 8,
    "USD": 9,
}

FOREIGN_CURRENCY_ORDER = ["USD", "GBP", "JPY", "CNY"]

TARGET_FONT = Font(name="Times New Roman", size=10)


@dataclass
class OcrLine:
    page: int
    y: float
    date_col: str
    detail_col: str
    deposit_col: str
    withdrawal_col: str
    balance_col: str


@dataclass
class StatementEntry:
    value_date: dt.date
    description: str
    currency: str
    amount: float
    source_page: int = 0
    source_y: float = 0.0


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def parse_number(value: str, is_jpy: bool = False) -> float | None:
    text = (value or "").replace("O", "0")
    if is_jpy:
        # JPY has no decimals. Strip ALL non-digit characters (except minus) to handle any OCR noise/dots.
        text = re.sub(r"[^-\d]", "", text)
        match = re.search(r"-?\d+", text)
        if not match:
            return None
        return float(match.group(0))
    
    match = NUMBER_RE.search(text)
    if not match:
        return None
    return float(match.group(0).replace(",", ""))


def is_cash_rebate_text(value: str) -> bool:
    flat = normalize_spaces(value).upper().replace(" ", "").replace("=", "E").replace("S", "5")
    return "CA5HREBATE" in flat or "CASHREBATE" in flat


def looks_like_purchase_detail(value: str) -> bool:
    text = normalize_spaces(value).upper()
    return bool(re.match(r"^(POS|MDC|MDG|MDO|VDC|TO PAYME|CREDIT INTEREST|CREDITINTEREST)\b", text))


def is_non_transaction_detail(value: str) -> bool:
    text = normalize_spaces(value).upper()
    compact = text.replace(" ", "")
    if not text:
        return False
    if text in {"HSBC", "BRANCH", "TRANSACTION DETAILS", "DATE"}:
        return True
    if "HAY WAH BUILDING" in text:
        return True
    if re.search(r"\d{3}-\d{6}-\d{3}", text):
        return True
    if "HONGKONGANDSHANGHAIBANKING" in compact:
        return True
    return False


def next_missing_foreign_currency(opening_balances: dict[str, float], current_currency: str | None) -> str | None:
    if not current_currency:
        return None
    try:
        start = FOREIGN_CURRENCY_ORDER.index(current_currency) + 1
    except ValueError:
        start = 0
    for currency in FOREIGN_CURRENCY_ORDER[start:]:
        if currency not in opening_balances:
            return currency
    return None


def looks_like_transaction_header(value: str) -> bool:
    text = normalize_spaces(value).upper()
    compact = re.sub(r"[^A-Z]", "", text)
    detail_like = "DETAIL" in compact or "DETAI" in compact or "ETAIL" in compact or "BETAIL" in compact
    return "DATE" in compact and ("TRANSACTION" in compact or "TRANSA" in compact) and detail_like


def clean_detail(value: str) -> str:
    text = normalize_spaces(value)
    if not text:
        return ""
    tokens = []
    for token in text.split():
        # Drop obvious Chinese OCR helper tokens but keep latin text.
        if re.search(r"[\u4e00-\u9fff]", token):
            continue
        tokens.append(token)
    return normalize_spaces(" ".join(tokens))


def humanize_description(value: str) -> str:
    text = normalize_spaces(value)
    if not text:
        return text

    # Remove embedded statement-date fragments like (21FEB26).
    text = re.sub(r"\(\d{1,2}[A-Z]{3}\d{2}\)", "", text, flags=re.IGNORECASE)
    text = normalize_spaces(text)

    # Simplify noisy provider suffixes to match manual-sheet readability.
    text = re.sub(r"\*?TRIP\s*HELP\.UBER\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\*?TRIPHELP\.UBER\b", "", text, flags=re.IGNORECASE)
    text = normalize_spaces(text)

    # Normalize common POS prefixes.
    text = re.sub(r"^POSMDC\b", "POS MDC", text, flags=re.IGNORECASE)
    text = re.sub(r"^POS\s+MDC\s*S\b", "POS MDC", text, flags=re.IGNORECASE)

    # Make POS lines more readable: "POS MDC ... / MERCHANT" -> "POS MDC - MERCHANT"
    if re.match(r"^POS\s+MDC\b", text, flags=re.IGNORECASE) and " / " in text:
        left, right = text.split(" / ", 1)
        right = normalize_spaces(right)
        if right:
            text = f"POS MDC - {right}"
        else:
            text = normalize_spaces(left)

    # Reader friendly cleanup for POS
    if text.upper().startswith("POS "):
        # Remove "MDC -" or "MDC "
        text = re.sub(r"^POS\s+MDC\s*(-\s*)?", "POS ", text, flags=re.IGNORECASE)
        
        merchant = text[4:].strip()
        # Clean up Uber specifically (lots of variants)
        if "UBER" in merchant.upper():
            text = "POS Uber"
        elif merchant.isupper():
            # Convert all-caps merchants to Title Case
            text = f"POS {merchant.title()}"

    text = re.sub(r"\bCASH\s*R[=\-]?BATE\b", "CASH REBATE", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCASHREBATE\b", "CASH REBATE", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCREDITINTEREST\b", "CREDIT INTEREST", text, flags=re.IGNORECASE)

    return normalize_spaces(text)


def parse_day_month(value: str, year: int) -> dt.date | None:
    compact = re.sub(r"[^0-9A-Za-z]", "", value.replace(" ", ""))
    numeric_october = re.match(r"^(\d{1,2})(?:0|O|Q)6$", compact, flags=re.IGNORECASE)
    if numeric_october:
        return dt.date(year, 10, int(numeric_october.group(1)))
    match = DATE_RE.match(compact) or DATE_RE.match(value)
    if not match:
        return None
    day = int(match.group(1))
    month_text = match.group(2).upper().replace("0", "O").replace("Q", "O")
    if month_text in {"OCF", "OCT"}:
        month_text = "OCT"
    month = MONTHS.get(month_text)
    if not month:
        return None
    return dt.date(year, month, day)


def parse_full_date(value: str) -> dt.date | None:
    match = FULL_DATE_RE.match(normalize_spaces(value))
    if not match:
        return None
    day = int(match.group(1))
    month = MONTHS.get(match.group(2).upper())
    if not month:
        return None
    year_text = match.group(3)
    year = int(year_text)
    if year < 100:
        year += 2000
    return dt.date(year, month, day)


def split_leading_day_month(value: str, year: int) -> tuple[dt.date | None, str]:
    text = normalize_spaces(value)
    match = re.match(r"^(\d{1,2}\s*[A-Za-z]{3})\b\s*(.*)$", text)
    if not match:
        return None, text
    parsed = parse_day_month(match.group(1), year)
    if not parsed:
        return None, text
    return parsed, normalize_spaces(match.group(2))


def extract_statement_year(lines: Iterable[OcrLine]) -> int:
    for line in lines:
        joined = " ".join(
            [line.date_col, line.detail_col, line.deposit_col, line.withdrawal_col, line.balance_col]
        )
        for match in STATEMENT_DATE_RE.finditer(joined):
            month_text = match.group(2)[:3].upper()
            if month_text in MONTHS:
                return int(match.group(3))
    return dt.date.today().year


def ocr_pdf_to_lines(pdf_path: Path, scale_dpi: int = 300) -> list[OcrLine]:
    """OCR entry point: Azure DI primary, RapidOCR fallback."""
    # Try Azure Document Intelligence first
    try:
        from azure_ocr_backend import ocr_pdf_to_lines as azure_ocr
        lines = azure_ocr(pdf_path, prefer_azure=True)
        if lines:
            return lines
    except Exception:
        pass

    # Fallback to local RapidOCR
    if os.environ.get("HSBC_DISABLE_RAPIDOCR", "").strip() == "1":
        return []

    try:
        from rapidocr_onnxruntime import RapidOCR
        import pypdfium2 as pdfium
    except Exception:
        return []

    doc = pdfium.PdfDocument(str(pdf_path))
    ocr = RapidOCR()
    lines: list[OcrLine] = []
    scale_dpi = 450
    # Base boundaries were established at 300 DPI. Scale them according to current DPI.
    dpi_scale = scale_dpi / 300.0
    
    x_date_max = 330 * dpi_scale
    x_detail_max = 1360 * dpi_scale
    x_deposit_max = 1680 * dpi_scale
    x_withdrawal_max = 1940 * dpi_scale

    for page_idx in range(len(doc)):
        import numpy as np
        pil_image = doc[page_idx].render(scale=scale_dpi / 72).to_pil()
        # Convert PIL image to numpy array for RapidOCR
        img_array = np.array(pil_image)
        result, _ = ocr(img_array)
        points = []
        for box, text, _confidence in (result or []):
            x = min(point[0] for point in box)
            y = min(point[1] for point in box)
            points.append((y, x, text))
        points.sort()

        grouped: list[tuple[float, list[tuple[float, str]]]] = []
        for y, x, text in points:
            if not grouped or abs(y - grouped[-1][0]) > 25:
                grouped.append((y, [(x, text)]))
            else:
                grouped[-1][1].append((x, text))

        for y, cells in grouped:
            ordered = sorted(cells)
            date_col = " ".join(text for x, text in ordered if x < x_date_max)
            detail_col = " ".join(text for x, text in ordered if x_date_max <= x < x_detail_max)
            deposit_col = " ".join(text for x, text in ordered if x_detail_max <= x < x_deposit_max)
            withdrawal_col = " ".join(text for x, text in ordered if x_deposit_max <= x < x_withdrawal_max)
            balance_col = " ".join(text for x, text in ordered if x_withdrawal_max <= x)
            lines.append(
                OcrLine(
                    page=page_idx + 1,
                    y=y,
                    date_col=normalize_spaces(date_col),
                    detail_col=normalize_spaces(detail_col),
                    deposit_col=normalize_spaces(deposit_col),
                    withdrawal_col=normalize_spaces(withdrawal_col),
                    balance_col=normalize_spaces(balance_col),
                )
            )

    return lines


def parse_hkd_entries(lines: list[OcrLine], year: int) -> tuple[list[StatementEntry], float | None, dt.date | None]:
    entries: list[StatementEntry] = []
    opening_balance: float | None = None
    opening_date: dt.date | None = None

    in_hkd_table = False
    saw_hkd_date_header = False
    current_date: dt.date | None = None
    current_desc_parts: list[str] = []

    for line in lines:
        row_date: dt.date | None = None
        joined = " ".join(
            [line.date_col, line.detail_col, line.deposit_col, line.withdrawal_col, line.balance_col]
        )
        joined_upper = joined.upper()
        # Handle common OCR errors like "Foreicn" or "Forging"
        if re.search(r"FOR.*CURR.*SAV", joined_upper, re.IGNORECASE):
            break

        if looks_like_transaction_header(joined) and "CCYDATE" not in joined_upper.replace(" ", ""):
            saw_hkd_date_header = True
            if "WITHDRAWAL" in joined_upper or "BALANCE" in joined_upper:
                in_hkd_table = True
                current_desc_parts = []
            continue
        if saw_hkd_date_header and ("WITHDRAWAL" in joined_upper or "BALANCE" in joined_upper):
            in_hkd_table = True
            current_desc_parts = []
            continue

        if not in_hkd_table:
            continue

        row_date = parse_day_month(line.date_col, year)

        detail = clean_detail(line.detail_col)
        if is_non_transaction_detail(detail):
            detail = ""
        if not row_date and detail:
            embedded_date, stripped_detail = split_leading_day_month(detail, year)
            if embedded_date:
                row_date = embedded_date
                detail = stripped_detail

        if detail and "B/F BALANCE" in detail.upper():
            opening_balance = parse_number(line.balance_col, is_jpy=False)
            opening_date = row_date or current_date
            if row_date:
                current_date = row_date
            current_desc_parts = []
            continue

        if row_date and detail and current_desc_parts and is_cash_rebate_text(" ".join(current_desc_parts)) and looks_like_purchase_detail(detail):
            current_desc_parts = []
        elif row_date and not current_desc_parts:
            current_desc_parts = []

        if detail:
            if current_desc_parts and is_cash_rebate_text(" ".join(current_desc_parts)) and looks_like_purchase_detail(detail):
                current_desc_parts = []
            current_desc_parts.append(detail)

        deposit = parse_number(line.deposit_col, is_jpy=False)
        withdrawal = parse_number(line.withdrawal_col, is_jpy=False)
        amount: float | None = None
        if deposit is not None:
            amount = deposit
        if withdrawal is not None:
            amount = -withdrawal

        # The first visible balance after B/F is treated as opening HKD balance.
        if opening_balance is None:
            maybe_balance = parse_number(line.balance_col, is_jpy=False)
            if maybe_balance is not None and amount is None:
                opening_balance = maybe_balance
                opening_date = current_date
        if amount is not None:
            if current_date and current_desc_parts:
                # If we have a new date on this line BUT no new description, the amount
                # likely belongs to the PREVIOUS date's transaction (delayed amount).
                # If we HAVE a description on this line, use the new date.
                entry_date = current_date if (row_date and not detail) else (row_date or current_date)
                
                description = normalize_spaces(" / ".join(current_desc_parts))
                description = humanize_description(description)
                entries.append(
                    StatementEntry(
                        value_date=entry_date,
                        description=description,
                        currency="HKD",
                        amount=amount,
                        source_page=line.page,
                        source_y=line.y,
                    )
                )
            current_desc_parts = []

        if row_date:
            current_date = row_date

    return entries, opening_balance, opening_date


def parse_foreign_entries(lines: list[OcrLine], year: int) -> tuple[list[StatementEntry], dict[str, float], dict[str, dt.date]]:
    entries: list[StatementEntry] = []
    opening_balances: dict[str, float] = {}
    opening_dates: dict[str, dt.date] = {}

    in_foreign_table = False
    current_currency: str | None = None
    current_date: dt.date | None = None
    current_desc_parts: list[str] = []
    last_balances: dict[str, float] = {}
    pending_known_amounts: dict[str, float] = {}
    pending_known_entries: dict[str, list[StatementEntry]] = {}
    pending_cash_rebates: dict[str, list[StatementEntry]] = {}

    for line in lines:
        row_date: dt.date | None = None
        explicit_currency: str | None = None
        joined = " ".join(
            [line.date_col, line.detail_col, line.deposit_col, line.withdrawal_col, line.balance_col]
        )
        joined_upper = joined.upper()
        if re.search(r"FOR.*CURR.*SAV", joined_upper, re.IGNORECASE):
            in_foreign_table = True
            current_desc_parts = []
            continue

        if not in_foreign_table:
            continue

        if re.search(r"TOTALRELATIONSHIPBALANCE", joined_upper.replace(" ", ""), re.IGNORECASE) or "TOTALRELATION" in joined_upper.replace(" ", ""):
            break

        # Clean header noise (e.g. dots in "CCY . Date")
        clean_header = joined_upper.replace(" ", "").replace(".", "").replace(":", "")
        if "CCYDATE" in clean_header or looks_like_transaction_header(joined):
            continue

        date_col_upper = line.date_col.upper()
        ccy_match = DATE_WITH_CCY_RE.match(date_col_upper.replace(" ", "")) or DATE_WITH_CCY_RE.match(date_col_upper)
        
        # Fallback: OCR sometimes misreads JPY as GBP in headers.
        # If we see "JPY" in the line but detected another currency, or if currency detection failed, trust the presence of JPY.
        is_jpy_described = "JPY" in joined_upper or "Jpy" in joined or "Õ¡ÿ" in joined
        
        if ccy_match:
            currency = ccy_match.group(1)
            if currency in CURRENCY_COLUMN:
                explicit_currency = currency
            # If the header regex matched something else (like GBP) but the line clearly mentions JPY, assume JPY.
            if is_jpy_described:
                explicit_currency = "JPY"
            if explicit_currency:
                current_currency = explicit_currency
                pending_known_amounts.setdefault(current_currency, 0.0)
                pending_known_entries.setdefault(current_currency, [])
                pending_cash_rebates.setdefault(current_currency, [])
            
            suffix_date = ccy_match.group(2)
            parsed = parse_day_month(suffix_date or "", year)
            if parsed:
                row_date = parsed
            current_desc_parts = []
        elif is_jpy_described and not current_currency:
            # If we don't have a currency yet but the line says JPY, treat it as the JPY section.
            current_currency = "JPY"
        else:
            only_date = parse_day_month(line.date_col, year)
            if only_date:
                row_date = only_date
                current_desc_parts = []

        if not current_currency:
            continue

        detail = clean_detail(line.detail_col)
        if is_non_transaction_detail(detail):
            detail = ""

        if detail and not row_date:
            embedded_date, stripped_detail = split_leading_day_month(detail, year)
            if embedded_date:
                row_date = embedded_date
                detail = stripped_detail

        if detail and "B/F BALANCE" in detail.upper():
            if not explicit_currency:
                inferred_currency = next_missing_foreign_currency(opening_balances, current_currency)
                if inferred_currency:
                    current_currency = inferred_currency
            is_jpy = current_currency == "JPY"
            opening = parse_number(line.balance_col, is_jpy=is_jpy)
            if opening is not None:
                opening_balances[current_currency] = opening
                last_balances[current_currency] = opening
                pending_known_amounts[current_currency] = 0.0
                pending_known_entries[current_currency] = []
                pending_cash_rebates[current_currency] = []
            opening_row_date = row_date or current_date
            if opening_row_date:
                opening_dates[current_currency] = opening_row_date
                current_date = opening_row_date
            current_desc_parts = []
            continue

        if detail:
            if current_desc_parts and is_cash_rebate_text(" ".join(current_desc_parts)) and looks_like_purchase_detail(detail):
                current_desc_parts = []
            current_desc_parts.append(detail)

        is_jpy = current_currency == "JPY"
        deposit = parse_number(line.deposit_col, is_jpy=is_jpy)
        withdrawal = parse_number(line.withdrawal_col, is_jpy=is_jpy)
        amount: float | None = None
        if deposit is not None:
            amount = deposit
        if withdrawal is not None:
            amount = -withdrawal
        entry_date = current_date if (row_date and not detail) else (row_date or current_date)
        if amount is None and current_currency and entry_date and detail and is_cash_rebate_text(detail):
            pending_cash_rebates.setdefault(current_currency, []).append(
                StatementEntry(
                    value_date=entry_date,
                    description=humanize_description(detail),
                    currency=current_currency,
                    amount=0.0,
                    source_page=line.page,
                    source_y=line.y,
                )
            )
        if amount is not None:
            if current_currency and entry_date and current_desc_parts:
                description = normalize_spaces(" / ".join(current_desc_parts))
                description = humanize_description(description)
                entry = StatementEntry(
                    value_date=entry_date,
                    description=description,
                    currency=current_currency,
                    amount=amount,
                    source_page=line.page,
                    source_y=line.y,
                )
                entries.append(entry)
                pending_known_amounts[current_currency] = pending_known_amounts.get(current_currency, 0.0) + amount
                pending_known_entries.setdefault(current_currency, []).append(entry)
            current_desc_parts = []

        balance = parse_number(line.balance_col, is_jpy=is_jpy)
        if current_currency and balance is not None and current_currency in last_balances:
            expected = last_balances[current_currency] + pending_known_amounts.get(current_currency, 0.0)
            missing_amount = balance - expected
            if is_jpy:
                missing_amount = round(missing_amount)
                tolerance = 0.1
            else:
                missing_amount = round(missing_amount, 2)
                tolerance = 0.01
            candidates = pending_cash_rebates.get(current_currency, [])
            if candidates and missing_amount > tolerance:
                allocations: list[float] = []
                if len(candidates) > 1 and is_jpy:
                    known_entries = pending_known_entries.get(current_currency, [])
                    for index, candidate in enumerate(candidates):
                        next_candidate_y = candidates[index + 1].source_y if index + 1 < len(candidates) else float("inf")
                        following_withdrawal = next(
                            (
                                abs(entry.amount)
                                for entry in known_entries
                                if entry.amount < 0 and candidate.source_y < entry.source_y < next_candidate_y
                            ),
                            None,
                        )
                        if following_withdrawal is None:
                            break
                        allocations.append(max(1.0, float(round(following_withdrawal * 0.005))))
                    if len(allocations) != len(candidates) or round(sum(allocations)) != missing_amount:
                        allocations = []
                if not allocations:
                    allocations = [float(missing_amount)]
                    candidates = [candidates[-1]]
                for recovered, recovered_amount in zip(candidates, allocations):
                    recovered.amount = recovered_amount
                    entries.append(recovered)
            last_balances[current_currency] = balance
            pending_known_amounts[current_currency] = 0.0
            pending_known_entries[current_currency] = []
            pending_cash_rebates[current_currency] = []

        if row_date:
            current_date = row_date

    return entries, opening_balances, opening_dates


def parse_uk_premier_entries(lines: list[OcrLine]) -> tuple[list[StatementEntry], float | None, dt.date | None]:
    entries: list[StatementEntry] = []
    opening_balance: float | None = None
    opening_date: dt.date | None = None
    in_table = False

    for line in lines:
        joined = " ".join([line.date_col, line.detail_col, line.deposit_col, line.withdrawal_col, line.balance_col])
        joined_upper = joined.upper()

        if "PAYMENT TYPE AND DETAILS" in joined_upper and "PAID OUT" in joined_upper:
            in_table = True
            continue

        if not in_table:
            continue

        date_value = parse_full_date(line.date_col)
        detail = humanize_description(clean_detail(line.detail_col))

        if date_value and "BALANCE BROUGHT FORWARD" in detail.upper():
            opening_balance = parse_number(line.balance_col)
            opening_date = date_value
            continue

        if date_value and "BALANCE CARRIED FORWARD" in detail.upper():
            break

        if not date_value or not detail:
            continue

        paid_out = parse_number(line.deposit_col)
        paid_in = parse_number(line.withdrawal_col)
        amount: float | None = None
        if paid_out is not None:
            amount = -paid_out
        elif paid_in is not None:
            amount = paid_in

        if amount is None:
            continue

        entries.append(
            StatementEntry(
                value_date=date_value,
                description=detail,
                currency="GBP",
                amount=amount,
                source_page=line.page,
                source_y=line.y,
            )
        )

    return entries, opening_balance, opening_date


def clear_sheet_data(ws) -> None:
    # Keep header rows (1-4), clear data rows in A:I.
    for row in range(5, ws.max_row + 1):
        for col in range(1, 10):
            ws.cell(row=row, column=col).value = None


def is_cash_rebate_entry(entry: StatementEntry) -> bool:
    desc = entry.description.upper().replace(" ", "")
    return "CASH" in desc and "REBATE" in desc and entry.amount > 0


def format_percentage(value: float) -> str:
    rounded = round(value, 1)
    text = f"{rounded:.1f}".rstrip("0").rstrip(".")
    return text


def annotate_cash_rebate_entries(entries: list[StatementEntry]) -> None:
    for i, entry in enumerate(entries):
        if not is_cash_rebate_entry(entry):
            continue

        reference_debit: StatementEntry | None = None

        # Prefer matching the next debit on the same date/currency.
        for j in range(i + 1, len(entries)):
            candidate = entries[j]
            if candidate.value_date != entry.value_date or candidate.currency != entry.currency:
                if candidate.value_date != entry.value_date:
                    break
                continue
            if candidate.amount < 0:
                reference_debit = candidate
                break
            if is_cash_rebate_entry(candidate):
                break

        # Fallback: nearest previous debit on same date/currency.
        if reference_debit is None:
            for j in range(i - 1, -1, -1):
                candidate = entries[j]
                if candidate.value_date != entry.value_date or candidate.currency != entry.currency:
                    continue
                if candidate.amount < 0:
                    reference_debit = candidate
                    break

        if reference_debit and reference_debit.amount != 0:
            pct = abs(entry.amount) / abs(reference_debit.amount) * 100
            entry.description = f"cash rebate - {format_percentage(pct)}%"
        else:
            entry.description = "cash rebate"


def ensure_target_sheet(workbook, template_sheet_name: str, target_sheet_name: str):
    template_sheet = workbook[template_sheet_name]
    if target_sheet_name in workbook.sheetnames:
        ws = workbook[target_sheet_name]
    else:
        ws = workbook.copy_worksheet(template_sheet)
        ws.title = target_sheet_name
    return ws


def write_entries_to_sheet(
    ws,
    entries: list[StatementEntry],
    opening_balances: dict[str, float],
    opening_date: dt.date | None,
) -> None:
    opening_style_by_col = {col: copy(ws.cell(row=5, column=col)._style) for col in range(1, 10)}
    txn_style_by_col = {col: copy(ws.cell(row=6, column=col)._style) for col in range(1, 10)}

    clear_sheet_data(ws)

    date_number_format = ws.cell(row=5, column=2).number_format or "mm/dd/yy;@"
    amount_number_formats = {
        col: ws.cell(row=5, column=col).number_format or "#,##0.00"
        for col in CURRENCY_COLUMN.values()
    }

    ws.cell(row=5, column=3).value = "Balance B/F"
    for col in range(1, 10):
        ws.cell(row=5, column=col)._style = copy(opening_style_by_col[col])
        ws.cell(row=5, column=col).font = copy(TARGET_FONT)
    if opening_date:
        ws.cell(row=5, column=2).value = dt.datetime.combine(opening_date, dt.time.min)
        ws.cell(row=5, column=2).number_format = date_number_format

    for currency, value in opening_balances.items():
        col = CURRENCY_COLUMN[currency]
        ws.cell(row=5, column=col).value = value
        ws.cell(row=5, column=col).number_format = amount_number_formats[col]

    row = 6
    previous_date: dt.date | None = None
    for entry in entries:
        for col in range(1, 10):
            ws.cell(row=row, column=col)._style = copy(txn_style_by_col[col])
            ws.cell(row=row, column=col).font = copy(TARGET_FONT)
        # Keep visual style close to the manually-entered sheets: only show date at first row per day.
        if previous_date != entry.value_date:
            ws.cell(row=row, column=2).value = dt.datetime.combine(entry.value_date, dt.time.min)
            ws.cell(row=row, column=2).number_format = date_number_format
            previous_date = entry.value_date
        else:
            ws.cell(row=row, column=2).value = None
        ws.cell(row=row, column=3).value = entry.description
        amount_col = CURRENCY_COLUMN[entry.currency]
        ws.cell(row=row, column=amount_col).value = entry.amount
        ws.cell(row=row, column=amount_col).number_format = amount_number_formats[amount_col]
        row += 1

    # Final font normalization pass so any remaining template formulas/labels in A:I
    # also match the required workbook font.
    for r in range(1, ws.max_row + 1):
        for c in range(1, 10):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).font = copy(TARGET_FONT)


def write_audit_sheet(workbook, lines: list[OcrLine], entries: list[StatementEntry]) -> None:
    sheet_name = "OCR Review"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    ws.append(["section", "page", "y", "date_col", "detail_col", "deposit_col", "withdrawal_col", "balance_col"])
    for line in lines:
        joined = " ".join([line.date_col, line.detail_col, line.deposit_col, line.withdrawal_col, line.balance_col])
        if not joined.strip():
            continue
        ws.append(
            [
                "ocr_line",
                line.page,
                round(line.y, 1),
                line.date_col,
                line.detail_col,
                line.deposit_col,
                line.withdrawal_col,
                line.balance_col,
            ]
        )

    ws.append([])
    ws.append(["section", "date", "currency", "amount", "description", "source_page", "source_y"])
    for entry in entries:
        ws.append(
            [
                "parsed_entry",
                entry.value_date.isoformat(),
                entry.currency,
                entry.amount,
                entry.description,
                entry.source_page,
                round(entry.source_y, 1),
            ]
        )

    widths = [16, 12, 10, 24, 50, 18, 18, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width


def extract_pdf_text_for_general_parser(pdf_path: Path) -> str:
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(str(pdf_path))
    except Exception:
        return ""
    chunks: list[str] = []
    for page in reader.pages:
        chunks.append(page.extract_text() or "")
    return "\n".join(chunks)


def ocr_lines_to_text(lines: list[OcrLine]) -> str:
    rows: list[str] = []
    for line in sorted(lines, key=lambda item: (item.page, item.y)):
        text = " ".join(
            value for value in [line.date_col, line.detail_col, line.deposit_col, line.withdrawal_col, line.balance_col] if value
        ).strip()
        if text:
            rows.append(text)
    return "\n".join(rows)


def parse_with_general_parser(pdf_path: Path, lines: list[OcrLine]) -> tuple[list[StatementEntry], dict[str, float], dt.date | None, str]:
    candidates = [("pdf-text", extract_pdf_text_for_general_parser(pdf_path)), ("ocr-text", ocr_lines_to_text(lines))]
    best_entries: list[StatementEntry] = []
    best_opening: dict[str, float] = {}
    best_opening_date: dt.date | None = None
    best_source = ""
    best_score = -1

    for source, text in candidates:
        if len("".join(ch for ch in text if ch.isalnum())) < 80:
            continue
        result = parse_general_statement_text(text, pdf_path)
        mapped: list[StatementEntry] = []
        for index, tx in enumerate(result.transactions):
            try:
                value_date = dt.datetime.strptime(tx.transaction_date, "%Y-%m-%d").date()
            except Exception:
                continue
            currency = (tx.currency or "HKD").upper()
            if currency not in CURRENCY_COLUMN:
                currency = "HKD"
            mapped.append(
                StatementEntry(
                    value_date=value_date,
                    description=humanize_description(normalize_spaces(tx.description)),
                    currency=currency,
                    amount=float(tx.amount),
                    source_page=1,
                    source_y=float(index),
                )
            )

        opening_balances = {
            currency.upper(): float(value)
            for currency, value in (result.opening_balances or {}).items()
            if currency.upper() in CURRENCY_COLUMN
        }
        opening_date: dt.date | None = None
        if result.opening_balance_date:
            try:
                opening_date = dt.datetime.strptime(result.opening_balance_date, "%Y-%m-%d").date()
            except Exception:
                opening_date = None

        score = len(mapped) + (2 if opening_balances else 0)
        if score > best_score:
            best_score = score
            best_entries = mapped
            best_opening = opening_balances
            best_opening_date = opening_date
            best_source = source

    return best_entries, best_opening, best_opening_date, best_source


def convert_statement(
    pdf_path: Path,
    template_path: Path,
    output_path: Path,
    template_sheet_name: str | None = None,
    sheet_name: str = "Auto",
    order_mode: str = "strict_pdf",
) -> dict:
    # Primary path: Azure Document Intelligence
    try:
        from azure_di_pipeline import convert_statement_azure
        return convert_statement_azure(
            pdf_path=pdf_path,
            template_path=template_path,
            output_path=output_path,
            template_sheet_name=template_sheet_name,
            sheet_name=sheet_name,
        )
    except Exception:
        pass

    # Fallback to legacy OCR pipeline
    lines = ocr_pdf_to_lines(pdf_path)
    year = extract_statement_year(lines)

    hkd_entries, hkd_opening, hkd_opening_date = parse_hkd_entries(lines, year)
    foreign_entries, foreign_openings, foreign_opening_dates = parse_foreign_entries(lines, year)
    uk_entries, uk_opening, uk_opening_date = parse_uk_premier_entries(lines)

    opening_balances = dict(foreign_openings)
    if hkd_opening is not None:
        opening_balances["HKD"] = hkd_opening
    if uk_opening is not None and not opening_balances:
        opening_balances["GBP"] = uk_opening

    opening_date = hkd_opening_date or uk_opening_date
    if not opening_date and foreign_opening_dates:
        opening_date = min(foreign_opening_dates.values())

    all_entries = hkd_entries + foreign_entries + uk_entries
    parser_engine = "layout-ocr"

    general_entries, general_opening, general_opening_date, general_source = parse_with_general_parser(pdf_path, lines)
    current_score = len(all_entries) + (2 if opening_balances else 0)
    general_score = len(general_entries) + (2 if general_opening else 0)
    if general_score > 0 and (current_score < 8 or general_score > current_score + 3):
        all_entries = general_entries
        parser_engine = f"general-{general_source}"
        if general_opening:
            opening_balances = general_opening
        if general_opening_date:
            opening_date = general_opening_date

    if order_mode == "strict_pdf":
        all_entries = sorted(all_entries, key=lambda e: (e.source_page, e.source_y))
    annotate_cash_rebate_entries(all_entries)

    workbook = load_workbook(template_path)
    selected_template_sheet = template_sheet_name or workbook.sheetnames[0]
    ws = ensure_target_sheet(workbook, selected_template_sheet, sheet_name)
    write_entries_to_sheet(ws, all_entries, opening_balances, opening_date)
    write_audit_sheet(workbook, lines, all_entries)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "statement_year": year,
        "opening_balances": opening_balances,
        "entries_written": len(all_entries),
        "output_path": str(output_path),
        "sheet_name": sheet_name,
        "opening_date": opening_date.isoformat() if opening_date else None,
        "parser_engine": parser_engine,
        "ocr_lines_count": len(lines),
        "has_any_entries": bool(all_entries),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert scanned HSBC statement PDF into Excel rows.")
    parser.add_argument("pdf", type=Path, help="Path to scanned statement PDF")
    parser.add_argument("template", type=Path, help="Path to existing Excel workbook template")
    parser.add_argument("output", type=Path, help="Path for output workbook")
    parser.add_argument(
        "--template-sheet",
        default=None,
        help="Sheet to copy formatting from (default: first sheet in workbook)",
    )
    parser.add_argument(
        "--sheet-name",
        default="Auto",
        help="Output sheet name to write transactions into",
    )
    parser.add_argument(
        "--order-mode",
        default="strict_pdf",
        choices=["strict_pdf"],
        help="Transaction ordering mode. 'strict_pdf' preserves source PDF sequence.",
    )
    args = parser.parse_args()

    result = convert_statement(
        pdf_path=args.pdf,
        template_path=args.template,
        output_path=args.output,
        template_sheet_name=args.template_sheet,
        sheet_name=args.sheet_name,
        order_mode=args.order_mode,
    )

    print(f"Statement year: {result['statement_year']}")
    print(f"Opening balances: {result['opening_balances']}")
    print(f"Entries written: {result['entries_written']}")
    print(f"Saved: {result['output_path']}")


if __name__ == "__main__":
    main()
