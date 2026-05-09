from flask import Flask, render_template, request, send_file, jsonify
import os
from pathlib import Path
from werkzeug.utils import secure_filename
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import openpyxl
from datetime import datetime
import json

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Config
UPLOAD_FOLDER = Path('uploads')
OUTPUT_FOLDER = Path('outputs')
UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

# Load Azure credentials from env or config
AZURE_ENDPOINT = os.environ.get('AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT', '')
AZURE_KEY = os.environ.get('AZURE_DOCUMENT_INTELLIGENCE_KEY', '')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    statement_type = request.form.get('type', 'hsbc')
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith('.pdf'):
        return jsonify({'error': 'Only PDF files are supported'}), 400
    
    # Save uploaded file
    filename = secure_filename(file.filename)
    filepath = UPLOAD_FOLDER / filename
    file.save(filepath)
    
    try:
        # Process with Azure DI
        client = DocumentAnalysisClient(AZURE_ENDPOINT, AzureKeyCredential(AZURE_KEY))
        
        with open(filepath, 'rb') as f:
            poller = client.begin_analyze_document('prebuilt-layout', document=f)
            result = poller.result()
        
        # Extract transactions based on type
        if statement_type == 'hsbc':
            transactions = extract_hsbc_transactions(result)
            output_file = create_hsbc_excel(transactions, filename)
        else:
            transactions = extract_amex_transactions(result)
            output_file = create_amex_excel(transactions, filename)
        
        return jsonify({
            'success': True,
            'transactions': len(transactions),
            'download_url': f'/download/{output_file.name}'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        # Clean up uploaded file
        filepath.unlink(missing_ok=True)

@app.route('/download/<filename>')
def download_file(filename):
    filepath = OUTPUT_FOLDER / filename
    if not filepath.exists():
        return jsonify({'error': 'File not found'}), 404
    return send_file(filepath, as_attachment=True)

@app.route('/config', methods=['GET', 'POST'])
def config():
    if request.method == 'POST':
        data = request.json
        os.environ['AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT'] = data.get('endpoint', '')
        os.environ['AZURE_DOCUMENT_INTELLIGENCE_KEY'] = data.get('key', '')
        return jsonify({'success': True})
    
    return jsonify({
        'endpoint': AZURE_ENDPOINT,
        'key': '***' + AZURE_KEY[-4:] if AZURE_KEY else ''
    })

def extract_hsbc_transactions(result):
    """Extract HSBC transactions from Azure DI result"""
    transactions = []
    
    for table in result.tables:
        # Look for transaction tables
        if not table.cells:
            continue
            
        # Check if this looks like a transaction table
        headers = [cell.content.strip().lower() for cell in table.cells if cell.row_index == 0]
        if not any(h in str(headers) for h in ['date', 'deposit', 'withdrawal', 'balance']):
            continue
        
        # Extract rows
        for row_idx in range(1, len(table.cells) // len(headers) if headers else 0):
            row_cells = [c for c in table.cells if c.row_index == row_idx]
            if len(row_cells) >= 4:
                transactions.append({
                    'date': row_cells[0].content if len(row_cells) > 0 else '',
                    'description': row_cells[1].content if len(row_cells) > 1 else '',
                    'deposit': row_cells[2].content if len(row_cells) > 2 else '',
                    'withdrawal': row_cells[3].content if len(row_cells) > 3 else '',
                    'balance': row_cells[4].content if len(row_cells) > 4 else ''
                })
    
    return transactions

def extract_amex_transactions(result):
    """Extract AMex transactions from Azure DI result"""
    transactions = []
    
    for table in result.tables:
        if not table.cells:
            continue
        
        headers = [cell.content.strip().lower() for cell in table.cells if cell.row_index == 0]
        if not any(h in str(headers) for h in ['date', 'description', 'amount']):
            continue
        
        for row_idx in range(1, len(table.cells) // len(headers) if headers else 0):
            row_cells = [c for c in table.cells if c.row_index == row_idx]
            if len(row_cells) >= 3:
                transactions.append({
                    'date': row_cells[0].content if len(row_cells) > 0 else '',
                    'merchant': row_cells[1].content if len(row_cells) > 1 else '',
                    'foreign_spend': row_cells[2].content if len(row_cells) > 2 else '',
                    'amount_hkd': row_cells[3].content if len(row_cells) > 3 else ''
                })
    
    return transactions

def create_hsbc_excel(transactions, original_filename):
    """Create HSBC Excel output"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    
    # Headers
    headers = ['Date', 'Description', 'Deposit', 'Withdrawal', 'Balance']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row_idx, tx in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=tx.get('date', ''))
        ws.cell(row=row_idx, column=2, value=tx.get('description', ''))
        ws.cell(row=row_idx, column=3, value=tx.get('deposit', ''))
        ws.cell(row=row_idx, column=4, value=tx.get('withdrawal', ''))
        ws.cell(row=row_idx, column=5, value=tx.get('balance', ''))
    
    output_file = OUTPUT_FOLDER / f"HSBC_{Path(original_filename).stem}.xlsx"
    wb.save(output_file)
    return output_file

def create_amex_excel(transactions, original_filename):
    """Create AMex Excel output"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    
    headers = ['Date', 'Merchant', 'Foreign Spend', 'Amount HKD']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row_idx, tx in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=tx.get('date', ''))
        ws.cell(row=row_idx, column=2, value=tx.get('merchant', ''))
        ws.cell(row=row_idx, column=3, value=tx.get('foreign_spend', ''))
        ws.cell(row=row_idx, column=4, value=tx.get('amount_hkd', ''))
    
    output_file = OUTPUT_FOLDER / f"AMEX_{Path(original_filename).stem}.xlsx"
    wb.save(output_file)
    return output_file

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
