#!/usr/bin/env python3
"""Unified GUI for Carmen OCR — HSBC statements + AMex statements.

Cross-platform tkinter app. Works on Windows, macOS, Linux.
"""

from __future__ import annotations

import json
import os
import re
import sys
import threading
from datetime import date
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# ---------------------------------------------------------------------------
# Config / settings
# ---------------------------------------------------------------------------

if getattr(sys, "frozen", False):
    # Running as PyInstaller .exe — put config next to the .exe, NOT in Program Files
    APP_DIR = Path(sys.executable).resolve().parent
    BUNDLE_DIR = Path(getattr(sys, "_MEIPASS", APP_DIR))
else:
    APP_DIR = Path(__file__).resolve().parent
    BUNDLE_DIR = APP_DIR

# Store settings next to the .exe (user-writable) or in user's Documents as fallback
if os.access(APP_DIR, os.W_OK):
    CONFIG_DIR = APP_DIR
else:
    # Fallback to user's Documents folder if .exe location is not writable
    CONFIG_DIR = Path.home() / "Documents" / "CarmenOCR"
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)

SETTINGS_PATH = CONFIG_DIR / "carmen_settings.json"
AZURE_CONFIG_PATH = CONFIG_DIR / "azure_config.json"


def default_settings() -> dict:
    return {
        "hsbc_input_folder": str(CONFIG_DIR / "Statements" / "Incoming"),
        "hsbc_output_folder": str(CONFIG_DIR / "Statements" / "Output"),
        "hsbc_template": str(CONFIG_DIR / "Statements" / "Docs" / "Su - HSBC One #833 (PayMe).xlsx"),
        "hsbc_template_sheet": "26",
        "amex_workbook": str(CONFIG_DIR / "Amex" / "AMEX.xlsx"),
        "amex_pdf_folder": str(CONFIG_DIR / "Amex"),
    }


def load_settings() -> dict:
    if not SETTINGS_PATH.exists():
        settings = default_settings()
        save_settings(settings)
        return settings
    with SETTINGS_PATH.open("r", encoding="utf-8") as f:
        settings = json.load(f)
    defaults = default_settings()
    for key, value in defaults.items():
        settings.setdefault(key, value)
    return settings


def save_settings(settings: dict) -> None:
    SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    with SETTINGS_PATH.open("w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2)


def resolve_path(raw: str) -> Path:
    value = (raw or "").strip().replace("\\", "/")
    value = value.replace("{APP_DIR}", str(APP_DIR))
    value = value.replace("{BUNDLE_DIR}", str(BUNDLE_DIR))
    value = value.replace("{CONFIG_DIR}", str(CONFIG_DIR))
    return Path(value).expanduser()


# ---------------------------------------------------------------------------
# Azure config
# ---------------------------------------------------------------------------

def load_azure_config() -> dict:
    if AZURE_CONFIG_PATH.exists():
        return json.loads(AZURE_CONFIG_PATH.read_text())
    return {}


def save_azure_config(endpoint: str, key: str) -> None:
    AZURE_CONFIG_PATH.write_text(json.dumps({"endpoint": endpoint, "key": key}, indent=2))


def get_public_ip() -> str | None:
    """Get the machine's public IP for Azure firewall allowlisting."""
    import urllib.request
    try:
        with urllib.request.urlopen("https://api.ipify.org", timeout=5) as resp:
            return resp.read().decode().strip()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class CarmenApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Carmen OCR — Statement Processor")
        self.geometry("950x650")
        self.settings = load_settings()
        self._review_changes: dict[str, dict] = {}

        # Notebook (tabs)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # HSBC tab
        self.hsbc_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.hsbc_frame, text="HSBC Statements")
        self._build_hsbc_tab()

        # AMex tab
        self.amex_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.amex_frame, text="AMex Statements")
        self._build_amex_tab()

        # Review Queue tab
        self.review_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.review_frame, text="Review Queue")
        self._build_review_tab()

        # Settings tab
        self.settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_frame, text="Settings")
        self._build_settings_tab()

        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(self, textvariable=self.status_var, relief="sunken").pack(fill="x", side="bottom")

    # -----------------------------------------------------------------------
    # HSBC Tab
    # -----------------------------------------------------------------------
    def _build_hsbc_tab(self) -> None:
        frame = ttk.Frame(self.hsbc_frame, padding=10)
        frame.pack(fill="both", expand=True)

        # Input folder
        ttk.Label(frame, text="Input Folder (PDFs)").grid(row=0, column=0, sticky="w")
        self.hsbc_input_var = tk.StringVar(value=self.settings.get("hsbc_input_folder", ""))
        ttk.Entry(frame, textvariable=self.hsbc_input_var, width=70).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(frame, text="Browse", command=self._pick_hsbc_input).grid(row=0, column=2)

        # Output folder
        ttk.Label(frame, text="Output Folder (Excel)").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.hsbc_output_var = tk.StringVar(value=self.settings.get("hsbc_output_folder", ""))
        ttk.Entry(frame, textvariable=self.hsbc_output_var, width=70).grid(row=1, column=1, sticky="ew", padx=5, pady=(8, 0))
        ttk.Button(frame, text="Browse", command=self._pick_hsbc_output).grid(row=1, column=2, pady=(8, 0))

        # Template
        ttk.Label(frame, text="Template Workbook").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.hsbc_template_var = tk.StringVar(value=self.settings.get("hsbc_template", ""))
        ttk.Entry(frame, textvariable=self.hsbc_template_var, width=70).grid(row=2, column=1, sticky="ew", padx=5, pady=(8, 0))
        ttk.Button(frame, text="Browse", command=self._pick_hsbc_template).grid(row=2, column=2, pady=(8, 0))

        # Template sheet
        ttk.Label(frame, text="Template Sheet").grid(row=3, column=0, sticky="w", pady=(8, 0))
        self.hsbc_sheet_var = tk.StringVar(value=self.settings.get("hsbc_template_sheet", "26"))
        ttk.Entry(frame, textvariable=self.hsbc_sheet_var, width=10).grid(row=3, column=1, sticky="w", padx=5, pady=(8, 0))

        ttk.Separator(frame).grid(row=4, column=0, columnspan=3, sticky="ew", pady=12)

        # PDF list
        list_frame = ttk.Frame(frame)
        list_frame.grid(row=5, column=0, columnspan=3, sticky="nsew")
        ttk.Label(list_frame, text="Statement PDFs").pack(side="left")
        ttk.Button(list_frame, text="Refresh", command=self._refresh_hsbc_pdfs).pack(side="right")

        self.hsbc_pdf_list = tk.Listbox(frame, height=12)
        self.hsbc_pdf_list.grid(row=6, column=0, columnspan=3, sticky="nsew", pady=(8, 10))

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=7, column=0, columnspan=3, sticky="ew")
        ttk.Button(btn_frame, text="Process Selected PDF", command=self._process_hsbc_selected).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Process All PDFs", command=self._process_hsbc_all).pack(side="left")

        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(6, weight=1)

        self._refresh_hsbc_pdfs()

    def _pick_hsbc_input(self) -> None:
        chosen = filedialog.askdirectory(initialdir=self.hsbc_input_var.get() or str(APP_DIR))
        if chosen:
            self.hsbc_input_var.set(chosen)
            self._refresh_hsbc_pdfs()

    def _pick_hsbc_output(self) -> None:
        chosen = filedialog.askdirectory(initialdir=self.hsbc_output_var.get() or str(APP_DIR))
        if chosen:
            self.hsbc_output_var.set(chosen)

    def _pick_hsbc_template(self) -> None:
        chosen = filedialog.askopenfilename(
            initialdir=str(Path(self.hsbc_template_var.get()).parent if self.hsbc_template_var.get() else APP_DIR),
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if chosen:
            self.hsbc_template_var.set(chosen)

    def _refresh_hsbc_pdfs(self) -> None:
        folder = resolve_path(self.hsbc_input_var.get())
        self.hsbc_pdf_list.delete(0, tk.END)
        if not folder.exists():
            return
        files = sorted(folder.glob("*.pdf"), key=lambda p: p.stat().st_mtime, reverse=True)
        for pdf in files:
            self.hsbc_pdf_list.insert(tk.END, pdf.name)
        self.status_var.set(f"HSBC: {len(files)} PDF(s) found")

    def _process_hsbc_selected(self) -> None:
        sel = self.hsbc_pdf_list.curselection()
        if not sel:
            messagebox.showwarning("No selection", "Select a PDF first.")
            return
        pdf_name = self.hsbc_pdf_list.get(sel[0])
        pdf_path = resolve_path(self.hsbc_input_var.get()) / pdf_name
        self._run_hsbc_process([pdf_path])

    def _process_hsbc_all(self) -> None:
        folder = resolve_path(self.hsbc_input_var.get())
        if not folder.exists():
            messagebox.showerror("Error", "Input folder does not exist.")
            return
        pdfs = sorted(folder.glob("*.pdf"))
        if not pdfs:
            messagebox.showinfo("Info", "No PDFs found.")
            return
        if not messagebox.askyesno("Confirm", f"Process {len(pdfs)} PDF(s)?"):
            return
        self._run_hsbc_process(pdfs)

    def _run_hsbc_process(self, pdfs: list[Path]) -> None:
        def worker():
            from hsbc_statement_to_excel import convert_statement
            template = resolve_path(self.hsbc_template_var.get())
            output_folder = resolve_path(self.hsbc_output_var.get())
            template_sheet = self.hsbc_sheet_var.get().strip() or "26"

            results = []
            for pdf in pdfs:
                sheet_name = pdf.stem[:6]  # e.g. 202503
                sheet_name = f"{sheet_name[4:6]}-{sheet_name[:4]}"
                output = output_folder / f"HSBC_{pdf.stem}.xlsx"
                try:
                    result = convert_statement(
                        pdf_path=pdf,
                        template_path=template,
                        output_path=output,
                        template_sheet_name=template_sheet,
                        sheet_name=sheet_name,
                    )
                    results.append(f"✓ {pdf.name}: {result['entries_written']} entries")
                except Exception as exc:
                    results.append(f"✗ {pdf.name}: {exc}")
                self.after(0, lambda r=results[-1]: self.status_var.set(r))

            self.after(0, lambda: messagebox.showinfo("Done", "\n".join(results)))
            self.after(0, self._refresh_hsbc_pdfs)

        self.status_var.set("Processing...")
        threading.Thread(target=worker, daemon=True).start()

    # -----------------------------------------------------------------------
    # AMex Tab
    # -----------------------------------------------------------------------
    def _build_amex_tab(self) -> None:
        frame = ttk.Frame(self.amex_frame, padding=10)
        frame.pack(fill="both", expand=True)

        # AMex workbook
        ttk.Label(frame, text="AMEX Workbook").grid(row=0, column=0, sticky="w")
        self.amex_workbook_var = tk.StringVar(value=self.settings.get("amex_workbook", ""))
        ttk.Entry(frame, textvariable=self.amex_workbook_var, width=70).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(frame, text="Browse", command=self._pick_amex_workbook).grid(row=0, column=2)

        # AMex PDF folder
        ttk.Label(frame, text="AMex PDF Folder").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.amex_folder_var = tk.StringVar(value=self.settings.get("amex_pdf_folder", ""))
        ttk.Entry(frame, textvariable=self.amex_folder_var, width=70).grid(row=1, column=1, sticky="ew", padx=5, pady=(8, 0))
        ttk.Button(frame, text="Browse", command=self._pick_amex_folder).grid(row=1, column=2, pady=(8, 0))

        ttk.Separator(frame).grid(row=2, column=0, columnspan=3, sticky="ew", pady=12)

        # PDF list
        list_frame = ttk.Frame(frame)
        list_frame.grid(row=3, column=0, columnspan=3, sticky="nsew")
        ttk.Label(list_frame, text="AMex Statement PDFs").pack(side="left")
        ttk.Button(list_frame, text="Refresh", command=self._refresh_amex_pdfs).pack(side="right")

        self.amex_pdf_list = tk.Listbox(frame, height=12)
        self.amex_pdf_list.grid(row=4, column=0, columnspan=3, sticky="nsew", pady=(8, 10))

        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=5, column=0, columnspan=3, sticky="ew")
        ttk.Button(btn_frame, text="Process Selected PDF", command=self._process_amex_selected).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Process All PDFs", command=self._process_amex_all).pack(side="left")

        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(4, weight=1)

        self._refresh_amex_pdfs()

    def _pick_amex_workbook(self) -> None:
        chosen = filedialog.askopenfilename(
            initialdir=str(Path(self.amex_workbook_var.get()).parent if self.amex_workbook_var.get() else APP_DIR),
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if chosen:
            self.amex_workbook_var.set(chosen)

    def _pick_amex_folder(self) -> None:
        chosen = filedialog.askdirectory(initialdir=self.amex_folder_var.get() or str(APP_DIR))
        if chosen:
            self.amex_folder_var.set(chosen)
            self._refresh_amex_pdfs()

    def _refresh_amex_pdfs(self) -> None:
        folder = resolve_path(self.amex_folder_var.get())
        self.amex_pdf_list.delete(0, tk.END)
        if not folder.exists():
            return
        files = sorted(folder.glob("*.pdf"), key=lambda p: p.stat().st_mtime, reverse=True)
        for pdf in files:
            self.amex_pdf_list.insert(tk.END, pdf.name)
        self.status_var.set(f"AMex: {len(files)} PDF(s) found")

    def _process_amex_selected(self) -> None:
        sel = self.amex_pdf_list.curselection()
        if not sel:
            messagebox.showwarning("No selection", "Select a PDF first.")
            return
        pdf_name = self.amex_pdf_list.get(sel[0])
        pdf_path = resolve_path(self.amex_folder_var.get()) / pdf_name
        self._run_amex_process([pdf_path])

    def _process_amex_all(self) -> None:
        folder = resolve_path(self.amex_folder_var.get())
        if not folder.exists():
            messagebox.showerror("Error", "Folder does not exist.")
            return
        pdfs = sorted(folder.glob("*.pdf"))
        if not pdfs:
            messagebox.showinfo("Info", "No PDFs found.")
            return
        if not messagebox.askyesno("Confirm", f"Process {len(pdfs)} PDF(s)?"):
            return
        self._run_amex_process(pdfs)

    def _run_amex_process(self, pdfs: list[Path]) -> None:
        def worker():
            from amex_pipeline import convert_amex_statement
            workbook = resolve_path(self.amex_workbook_var.get())

            results = []
            for pdf in pdfs:
                try:
                    result = convert_amex_statement(
                        pdf_path=pdf,
                        workbook_path=workbook,
                    )
                    results.append(f"✓ {pdf.name}: {result['entries_written']} entries")
                except Exception as exc:
                    results.append(f"✗ {pdf.name}: {exc}")
                self.after(0, lambda r=results[-1]: self.status_var.set(r))

            self.after(0, lambda: messagebox.showinfo("Done", "\n".join(results)))
            self.after(0, self._refresh_amex_pdfs)

        self.status_var.set("Processing AMex...")
        threading.Thread(target=worker, daemon=True).start()

    # -----------------------------------------------------------------------
    # Settings Tab
    # -----------------------------------------------------------------------
    def _build_settings_tab(self) -> None:
        frame = ttk.Frame(self.settings_frame, padding=10)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Azure Document Intelligence Endpoint").grid(row=0, column=0, sticky="w")
        self.azure_endpoint_var = tk.StringVar(value=load_azure_config().get("endpoint", "https://documentbank.cognitiveservices.azure.com/"))
        ttk.Entry(frame, textvariable=self.azure_endpoint_var, width=70).grid(row=0, column=1, sticky="ew", padx=5)

        ttk.Label(frame, text="Azure Key").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.azure_key_var = tk.StringVar(value=load_azure_config().get("key", ""))
        ttk.Entry(frame, textvariable=self.azure_key_var, width=70, show="*").grid(row=1, column=1, sticky="ew", padx=5, pady=(8, 0))

        ttk.Button(frame, text="Save Azure Credentials", command=self._save_azure).grid(row=2, column=1, sticky="w", pady=(8, 0))
        ttk.Button(frame, text="Test Connection", command=self._test_azure_connection).grid(row=2, column=1, sticky="e", pady=(8, 0))

        ttk.Separator(frame).grid(row=3, column=0, columnspan=2, sticky="ew", pady=12)

        ttk.Button(frame, text="Save All Settings", command=self._save_all_settings).grid(row=4, column=1, sticky="w")

        frame.columnconfigure(1, weight=1)

    def _save_azure(self) -> None:
        endpoint = self.azure_endpoint_var.get().strip()
        key = self.azure_key_var.get().strip()
        if not endpoint or not key:
            messagebox.showwarning("Missing Credentials", "Both endpoint and key are required.")
            return
        if not endpoint.startswith(("https://", "http://")):
            messagebox.showwarning("Invalid Endpoint", "Endpoint must start with https://")
            return
        if len(key) < 20:
            if not messagebox.askyesno("Short Key", "The key looks unusually short. Continue anyway?"):
                return
        save_azure_config(endpoint, key)
        messagebox.showinfo("Saved", "Azure credentials saved securely.")
        # Clear the key from memory after saving
        self.azure_key_var.set("")
        self.azure_key_var.set(key)

    def _test_azure_connection(self) -> None:
        """Test Azure connection without processing a statement."""
        endpoint = self.azure_endpoint_var.get().strip()
        key = self.azure_key_var.get().strip()
        if not endpoint or not key:
            messagebox.showwarning("Missing Credentials", "Enter endpoint and key first.")
            return
        
        def worker():
            try:
                from azure.ai.formrecognizer import DocumentAnalysisClient
                from azure.core.credentials import AzureKeyCredential
                client = DocumentAnalysisClient(endpoint, AzureKeyCredential(key))
                # Just create client - don't make API call to avoid charges
                self.after(0, lambda: messagebox.showinfo("Success", "Azure credentials are valid.\nConnection ready."))
                self.after(0, lambda: self.status_var.set("Azure connection OK"))
            except Exception as exc:
                self.after(0, lambda: messagebox.showerror("Connection Failed", f"Could not connect to Azure:\n{exc}"))
                self.after(0, lambda: self.status_var.set("Azure connection failed"))
        
        self.status_var.set("Testing Azure connection...")
        threading.Thread(target=worker, daemon=True).start()

    # -----------------------------------------------------------------------
    # Review Queue Tab
    # -----------------------------------------------------------------------
    def _build_review_tab(self) -> None:
        frame = ttk.Frame(self.review_frame, padding=10)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Uncategorized Merchants — Select a row and use dropdowns to categorize").pack(anchor="w")

        # Treeview for Review Items
        columns = ("Date", "Merchant", "Amount", "Category", "Belongs To", "Statement", "Status")
        self.review_tree = ttk.Treeview(frame, columns=columns, show="headings", height=20)
        for col in columns:
            self.review_tree.heading(col, text=col)
            self.review_tree.column(col, width=120 if col in ("Date", "Amount", "Category", "Belongs To") else 180)
        self.review_tree.column("Amount", width=100, anchor="e")

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.review_tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.review_tree.xview)
        self.review_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.review_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Action buttons
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x", pady=(10, 0))
        ttk.Button(btn_frame, text="Refresh Queue", command=self._refresh_review_queue).pack(side="left", padx=(0, 8))
        ttk.Button(btn_frame, text="Save Changes", command=self._save_review_changes).pack(side="left", padx=(0, 8))
        ttk.Label(btn_frame, text="Double-click Category/Belongs To cell to edit").pack(side="right")

        # Bind double-click to edit
        self.review_tree.bind("<Double-1>", self._on_review_double_click)

        self._refresh_review_queue()

    def _refresh_review_queue(self) -> None:
        workbook = resolve_path(self.amex_workbook_var.get())
        if not workbook.exists():
            self.status_var.set("AMex workbook not found")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(workbook, data_only=True)
            ws = wb["Review Items"]

            self.review_tree.delete(*self.review_tree.get_children())

            for row in range(5, ws.max_row + 1):
                date_val = ws.cell(row=row, column=1).value
                merchant = ws.cell(row=row, column=2).value
                amount = ws.cell(row=row, column=3).value
                category = ws.cell(row=row, column=4).value
                belongs = ws.cell(row=row, column=5).value
                statement = ws.cell(row=row, column=6).value
                status = ws.cell(row=row, column=7).value

                if not merchant:
                    continue

                # Only show uncategorized or review items
                if category == "Uncategorised" or belongs == "Review":
                    date_str = date_val.strftime("%Y-%m-%d") if hasattr(date_val, "strftime") else str(date_val)[:10]
                    amount_str = f"{amount:,.2f}" if isinstance(amount, (int, float)) else str(amount)
                    self.review_tree.insert("", "end", iid=str(row), values=(
                        date_str, merchant, amount_str, category or "", belongs or "", statement or "", status or ""
                    ))

            self.status_var.set(f"Review Queue: {len(self.review_tree.get_children())} items pending")
        except Exception as exc:
            self.status_var.set(f"Error loading review queue: {exc}")

    def _on_review_double_click(self, event) -> None:
        region = self.review_tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        col = self.review_tree.identify_column(event.x)
        item = self.review_tree.identify_row(event.y)
        if not item:
            return

        # Only allow editing Category (col 4) or Belongs To (col 5)
        if col not in ("#4", "#5"):
            return

        values = self.review_tree.item(item, "values")
        current = values[3] if col == "#4" else values[4]

        # Create dropdown
        x, y, width, height = self.review_tree.bbox(item, col)
        var = tk.StringVar(value=current)

        if col == "#4":
            # Category dropdown
            combo = ttk.Combobox(self.review_frame, textvariable=var, values=self._get_categories(), state="readonly")
        else:
            # Belongs To dropdown
            combo = ttk.Combobox(self.review_frame, textvariable=var, values=["Family", "Payment", "Review", "Su"], state="readonly")

        combo.place(x=x, y=y, width=width, height=height)
        combo.focus()

        def on_select(event=None):
            new_val = var.get()
            self.review_tree.set(item, column=col, value=new_val)
            combo.destroy()
            # Mark as changed
            self._review_changes[item] = {
                "row": int(item),
                "category": self.review_tree.set(item, "Category") if col != "#4" else new_val,
                "belongs_to": self.review_tree.set(item, "Belongs To") if col != "#5" else new_val,
            }

        combo.bind("<<ComboboxSelected>>", on_select)
        combo.bind("<FocusOut>", lambda e: combo.destroy())

    def _get_categories(self) -> list[str]:
        workbook = resolve_path(self.amex_workbook_var.get())
        if not workbook.exists():
            return []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(workbook, data_only=True)
            ws = wb["Lists"]
            cats = []
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if val:
                    cats.append(str(val))
            return cats
        except Exception:
            return []

    def _save_review_changes(self) -> None:
        if not hasattr(self, "_review_changes") or not self._review_changes:
            messagebox.showinfo("No Changes", "Nothing to save.")
            return

        workbook = resolve_path(self.amex_workbook_var.get())
        try:
            from openpyxl import load_workbook
            wb = load_workbook(workbook)
            ws = wb["Review Items"]

            for item_id, change in self._review_changes.items():
                row = change["row"]
                ws.cell(row=row, column=4, value=change["category"])
                ws.cell(row=row, column=5, value=change["belongs_to"])

            wb.save(workbook)
            self._review_changes.clear()
            messagebox.showinfo("Saved", "Changes saved to Review Items.")
            self.status_var.set("Review changes saved")
            self._refresh_review_queue()
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to save: {exc}")
            self.status_var.set(f"Save error: {exc}")

    def _save_all_settings(self) -> None:
        settings = {
            "hsbc_input_folder": self.hsbc_input_var.get().strip(),
            "hsbc_output_folder": self.hsbc_output_var.get().strip(),
            "hsbc_template": self.hsbc_template_var.get().strip(),
            "hsbc_template_sheet": self.hsbc_sheet_var.get().strip(),
            "amex_workbook": self.amex_workbook_var.get().strip(),
            "amex_pdf_folder": self.amex_folder_var.get().strip(),
        }
        save_settings(settings)
        messagebox.showinfo("Saved", "Settings saved.")


def main() -> None:
    if sys.platform.startswith("win"):
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass
    app = CarmenApp()
    app.mainloop()


if __name__ == "__main__":
    main()
