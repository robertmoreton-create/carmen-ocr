#!/usr/bin/env python3
"""Create a per-statement test sheet in the AMEX workbook."""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

# Load the workbook with the extracted data
wb = load_workbook('Amex/AMEX_TEST3.xlsx')
ws_txn = wb['Amex Transactions']

# Find all rows for Amex Mar-2026
statement_name = 'Amex Mar-2026'
rows = []
for row in range(5, ws_txn.max_row + 1):
    if ws_txn.cell(row=row, column=1).value == statement_name:
        rows.append(row)

print(f"Found {len(rows)} rows for {statement_name}")

# Create new sheet
sheet_name = "Mar-2026-TEST"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(title=sheet_name)

# Headers
headers = ['Date', 'Merchant', 'Category', 'Belongs To', 'Foreign Spend', 'Amount HKD', 'Source Page']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Copy data
for i, src_row in enumerate(rows, 2):
    ws.cell(row=i, column=1, value=ws_txn.cell(row=src_row, column=2).value)  # Date
    ws.cell(row=i, column=2, value=ws_txn.cell(row=src_row, column=3).value)  # Merchant
    ws.cell(row=i, column=3, value=ws_txn.cell(row=src_row, column=4).value)  # Category
    ws.cell(row=i, column=4, value=ws_txn.cell(row=src_row, column=5).value)  # Belongs To
    ws.cell(row=i, column=5, value=ws_txn.cell(row=src_row, column=6).value)  # Foreign Spend
    ws.cell(row=i, column=6, value=ws_txn.cell(row=src_row, column=7).value)  # Amount HKD
    ws.cell(row=i, column=7, value=ws_txn.cell(row=src_row, column=8).value)  # Source Page

# Auto-adjust column widths
for col in range(1, 8):
    max_length = 0
    for row in range(1, len(rows) + 2):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[chr(64 + col)].width = min(max_length + 2, 50)

# Save
output_path = 'Amex/AMEX_WITH_TEST_SHEET.xlsx'
wb.save(output_path)
print(f"Saved to: {output_path}")
print(f"Sheet '{sheet_name}' created with {len(rows)} transactions")
