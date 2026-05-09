#!/usr/bin/env python3
"""American Express statement extraction via Azure Document Intelligence.

Extracts transactions from AMex PDFs and appends them to the AMEX.xlsx workbook.

Target sheet: 'Amex Transactions'
Columns: Statement | Date | Merchant | Category | Belongs To | Foreign Spend | Amount HKD | Source Page | Review

Also adds new merchants to 'Review Items' queue for categorization.
"""

from __future__ import annotations

import os
import re
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# Azure credentials (same fallback chain)
# ---------------------------------------------------------------------------

AZURE_ENDPOINT = os.environ.get(
    "AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT",
    "https://documentbank.cognitiveservices.azure.com/",
)
AZURE_KEY = os.environ.get("AZURE_DOCUMENT_INTELLIGENCE_KEY", "")

if not AZURE_KEY:
    _cfg_path = Path(__file__).with_name("azure_config.json")
    if _cfg_path.exists():
        import json
        _cfg = json.loads(_cfg_path.read_text())
        AZURE_ENDPOINT = _cfg.get("endpoint", AZURE_ENDPOINT)
        AZURE_KEY = _cfg.get("key", "")

# Production builds should NOT include test_azure.py —
# the key must come from env var or azure_config.json only.
# The dev fallback below is disabled by default; set CARMEN_DEV=1 to enable.
if not AZURE_KEY and os.environ.get("CARMEN_DEV") == "1":
    try:
        import importlib.util
        _test_azure_path = Path(__file__).parent / "test_azure.py"
        if _test_azure_path.exists():
            _spec = importlib.util.spec_from_file_location("test_azure", _test_azure_path)
            _mod = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)
            AZURE_KEY = getattr(_mod, "KEY", "")
            AZURE_ENDPOINT = getattr(_mod, "ENDPOINT", AZURE_ENDPOINT)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Types
# ---------------------------------------------------------------------------

@dataclass
class AmexTransaction:
    date: date
    merchant: str
    foreign_spend: str
    amount_hkd: float
    is_credit: bool
    source_page: int


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def _parse_statement_date(text: str) -> date | None:
    """Extract statement date like 'March 26, 2026' from text."""
    m = re.search(r"(\b[A-Za-z]+\s+\d{1,2},?\s+20\d{2}\b)", text)
    if m:
        try:
            return datetime.strptime(m.group(1).replace(",", ""), "%B %d %Y").date()
        except ValueError:
            pass
    # Try shorter format
    m = re.search(r"Statement\s+includes\s+.*by\s+([A-Za-z]+\s+\d{1,2},?\s+20\d{2})", text, re.IGNORECASE)
    if m:
        try:
            return datetime.strptime(m.group(1).replace(",", ""), "%B %d %Y").date()
        except ValueError:
            pass
    return None


def _parse_transaction_date(date_str: str, year_hint: int) -> date | None:
    """Parse dates like 'February 26' or 'March 1'."""
    text = _norm(date_str)
    if not text:
        return None
    # "February 26" or "Feb 26"
    m = re.match(r"([A-Za-z]+)\s+(\d{1,2})", text)
    if not m:
        return None
    month_str = m.group(1)
    day = int(m.group(2))
    # Resolve month
    month_map = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    month = month_map.get(month_str.lower())
    if not month:
        return None
    return date(year_hint, month, day)


def _extract_amount(amount_str: str) -> float | None:
    if not amount_str:
        return None
    cleaned = amount_str.replace(",", "").replace(" ", "").strip()
    m = re.search(r"-?\d+\.?\d*", cleaned)
    if not m:
        return None
    try:
        return float(m.group())
    except ValueError:
        return None


def _clean_merchant(merchant: str) -> str:
    """Clean up merchant names."""
    text = _norm(merchant)
    # Remove "All-inclusive exchange rate# : 0.051" suffixes
    text = re.sub(r"All-inclusive exchange rate#?\s*:?\s*\d+\.?\d*", "", text, flags=re.IGNORECASE)
    # Remove "meng tian jun" suffix (cardholder name on some transactions)
    text = re.sub(r"\bmeng tian jun\b", "", text, flags=re.IGNORECASE)
    # Remove location suffixes like "HOKKAIDO", "CENTRAL", "HONG KONG" if they appear at end
    text = re.sub(r"\s+(HOKKAIDO|HONG KONG|SINGAPORE|TOKYO|LONDON|CENTRAL|ADMIRALTY|KOWLOON|ISLANDS?)\s*$", "", text, flags=re.IGNORECASE)
    # Remove trailing "PG"
    text = re.sub(r"\s+PG\s*$", "", text)
    # Remove cardholder names like "LEE SU HWEI" etc
    text = re.sub(r"\b(LEE SU HWEI|SU HWEI LEE|LEE SH)\b", "", text, flags=re.IGNORECASE)
    # Clean up excessive spaces
    text = _norm(text)
    # Title case for readability
    if text.isupper():
        text = text.title()
    return text


def _is_date_only(value: str) -> bool:
    """True if the value looks like just a date (e.g. 'February 26', 'March 1')."""
    text = _norm(value)
    if not text:
        return False
    m = re.match(r"^[A-Za-z]+\s+\d{1,2}$", text)
    return bool(m)


# ---------------------------------------------------------------------------
# Azure extraction
# ---------------------------------------------------------------------------

def _azure_extract(pdf_path: Path) -> tuple[str, list[AmexTransaction]]:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential

    client = DocumentAnalysisClient(AZURE_ENDPOINT, AzureKeyCredential(AZURE_KEY))

    with pdf_path.open("rb") as f:
        poller = client.begin_analyze_document("prebuilt-layout", f)
        result = poller.result()

    # Extract statement date from paragraphs
    all_text = "\n".join([p.content for p in result.paragraphs])
    statement_date = _parse_statement_date(all_text)
    if not statement_date:
        # Fallback: extract from filename
        m = re.search(r"20(\d{2})(\d{2})(\d{2})", pdf_path.name)
        if m:
            statement_date = date(2000 + int(m.group(1)), int(m.group(2)), int(m.group(3)))
        else:
            statement_date = date.today()

    year_hint = statement_date.year

    transactions: list[AmexTransaction] = []

    for table in result.tables:
        # Check if this is a transaction table
        header = " ".join([c.content for c in table.cells if c.row_index == 0]).lower()
        if "details" not in header or "amount" not in header:
            continue

        # Find column indices from header row
        # AMex layout: Col 0 = Date, Col 1 = Merchant/Details, Col 2 = Foreign Spend, Col 3 = Amount
        # But header row says "DETAILS" in col 0 which is misleading — col 0 is actually dates
        date_idx = 0
        detail_idx = 1
        foreign_idx = 2
        amount_idx = 3

        for cell in table.cells:
            if cell.row_index != 0:
                continue
            text = cell.content.lower().strip()
            col = cell.column_index
            # Only override if we find explicit headers
            if text == "date":
                date_idx = col
            elif "foreign" in text or "spend" in text:
                foreign_idx = col
            elif "amount" in text or "hk" in text:
                amount_idx = col
            # "DETAILS" is in col 0 but that's the date column — don't use it for detail_idx

        if amount_idx is None:
            continue

        # Group cells by row
        rows: dict[int, dict[int, str]] = {}
        for cell in table.cells:
            if cell.row_index == 0:
                continue
            if cell.row_index not in rows:
                rows[cell.row_index] = {}
            rows[cell.row_index][cell.column_index] = cell.content.strip()

        current_date: date | None = None
        pending_merchant = ""

        for row_idx in sorted(rows.keys()):
            cells = rows[row_idx]
            date_val = cells.get(date_idx, "")
            detail_val = cells.get(detail_idx, "")
            foreign_val = cells.get(foreign_idx, "")
            amount_val = cells.get(amount_idx, "")

            # Parse date
            new_date = _parse_transaction_date(date_val, year_hint)
            if new_date:
                current_date = new_date

            # Skip header-like rows
            if "new transactions for" in detail_val.lower():
                continue
            if "all-inclusive exchange rate" in detail_val.lower() and not foreign_val:
                continue

            # Skip payment rows (handled separately or as credits)
            is_payment = "payment received" in detail_val.lower() or "autopay" in detail_val.lower()
            is_credit = "cr" in detail_val.lower() or is_payment

            # Amount
            amount = _extract_amount(amount_val)
            if amount is None:
                # This might be a continuation row — save merchant for next row
                if detail_val and not is_payment and not _is_date_only(detail_val):
                    pending_merchant = detail_val
                continue

            # Build merchant name from detail column
            merchant = detail_val
            if pending_merchant and merchant:
                # If current row has exchange rate info, use pending as merchant
                if "all-inclusive exchange rate" in merchant.lower():
                    merchant = pending_merchant
                elif pending_merchant.lower() not in merchant.lower():
                    merchant = f"{pending_merchant} {merchant}"
            pending_merchant = ""

            # Clean merchant
            merchant = _clean_merchant(merchant)

            # Skip if still no useful merchant
            if not merchant or merchant.lower() in ["cr", "", "all-inclusive exchange rate"]:
                continue

            # Skip if merchant is just a date (happens when date column bleeds into detail)
            if _is_date_only(merchant):
                continue

            # Determine credit (negative amount)
            if is_credit or is_payment:
                amount = -abs(amount)

            transactions.append(
                AmexTransaction(
                    date=current_date or statement_date,
                    merchant=merchant,
                    foreign_spend=_norm(foreign_val),
                    amount_hkd=amount,
                    is_credit=is_credit or amount < 0,
                    source_page=1,
                )
            )

    # Sort by date
    transactions.sort(key=lambda t: t.date)

    # Build statement name
    statement_name = f"Amex {statement_date.strftime('%b-%Y')}"

    return statement_name, transactions


# ---------------------------------------------------------------------------
# Excel operations
# ---------------------------------------------------------------------------

def _get_next_row(ws) -> int:
    """Find first empty row after existing data."""
    for row in range(5, ws.max_row + 2):
        if ws.cell(row=row, column=2).value is None:
            return row
    return ws.max_row + 1


def _merchant_exists_in_review(ws_review, merchant: str) -> bool:
    """Check if merchant already in Review Items."""
    merchant_clean = _norm(merchant).lower()
    for row in range(5, ws_review.max_row + 1):
        val = ws_review.cell(row=row, column=2).value
        if val and _norm(str(val)).lower() == merchant_clean:
            return True
    return False


def _lookup_category(wb, merchant: str) -> tuple[str | None, str | None]:
    """Look up merchant in Category sheet. Returns (category, belongs_to) or (None, None)."""
    if "Category" not in wb.sheetnames:
        return None, None
    ws = wb["Category"]
    merchant_clean = _norm(merchant).lower()

    # Try exact match first
    for row in range(2, ws.max_row + 1):
        match_text = ws.cell(row=row, column=1).value
        if match_text and _norm(str(match_text)).lower() == merchant_clean:
            cat = ws.cell(row=row, column=2).value
            belongs = ws.cell(row=row, column=3).value
            return str(cat) if cat else None, str(belongs) if belongs else None

    # Try partial match (merchant contains match text)
    for row in range(2, ws.max_row + 1):
        match_text = ws.cell(row=row, column=1).value
        if match_text:
            mt_clean = _norm(str(match_text)).lower()
            if mt_clean in merchant_clean or merchant_clean in mt_clean:
                cat = ws.cell(row=row, column=2).value
                belongs = ws.cell(row=row, column=3).value
                return str(cat) if cat else None, str(belongs) if belongs else None

    return None, None


def _add_to_review_items(wb, merchant: str, tx_date: date, amount: float, statement: str) -> None:
    """Add new merchant to Review Items queue if not already there."""
    if "Review Items" not in wb.sheetnames:
        return
    ws = wb["Review Items"]

    if _merchant_exists_in_review(ws, merchant):
        return

    # Auto-categorize if known
    category, belongs_to = _lookup_category(wb, merchant)

    # Find next empty row
    next_row = ws.max_row + 1

    ws.cell(row=next_row, column=1).value = datetime.combine(tx_date, datetime.min.time())
    ws.cell(row=next_row, column=2).value = merchant
    ws.cell(row=next_row, column=3).value = amount
    ws.cell(row=next_row, column=4).value = category or "Uncategorised"
    ws.cell(row=next_row, column=5).value = belongs_to or "Review"
    ws.cell(row=next_row, column=6).value = statement

    # Copy formula from row above for Status column
    if next_row > 5:
        above_formula = ws.cell(row=next_row - 1, column=7).value
        if above_formula and isinstance(above_formula, str) and above_formula.startswith("="):
            new_formula = re.sub(r"(\D)(\d+)", lambda m: m.group(1) + str(int(m.group(2)) + 1) if int(m.group(2)) >= 5 else m.group(0), above_formula)
            ws.cell(row=next_row, column=7).value = new_formula

    # Extend data validation to cover new row
    for dv in ws.data_validations.dataValidation:
        old_ranges = list(dv.sqref.ranges)
        new_ranges = []
        for rng in old_ranges:
            start, end = str(rng).split(":")
            col = re.match(r"([A-Z]+)", end).group(1)
            new_end = f"{col}{next_row}"
            new_ranges.append(f"{start}:{new_end}")
        dv.sqref = " ".join(new_ranges)


def _add_to_rule_updates(wb, merchant: str) -> None:
    """Add new merchant to Rule Updates if not already there."""
    if "Rule Updates" not in wb.sheetnames:
        return
    ws = wb["Rule Updates"]

    merchant_clean = _norm(merchant).lower()
    for row in range(5, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and _norm(str(val)).lower() == merchant_clean:
            return

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1).value = merchant
    ws.cell(row=next_row, column=2).value = "Uncategorised"
    ws.cell(row=next_row, column=3).value = "Review"

    # Copy formulas from above
    if next_row > 5:
        for col in [4, 5, 6]:
            above = ws.cell(row=next_row - 1, column=col).value
            if above and isinstance(above, str) and above.startswith("="):
                new_formula = re.sub(r"(\D)(\d+)", lambda m: m.group(1) + str(int(m.group(2)) + 1) if int(m.group(2)) >= 5 else m.group(0), above)
                ws.cell(row=next_row, column=col).value = new_formula


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def convert_amex_statement(
    pdf_path: Path,
    workbook_path: Path,
    output_path: Path | None = None,
) -> dict:
    """Extract AMex statement and append to AMEX.xlsx workbook."""

    statement_name, transactions = _azure_extract(pdf_path)

    if not transactions:
        return {
            "statement_name": statement_name,
            "entries_written": 0,
            "output_path": str(output_path or workbook_path),
            "error": "No transactions extracted",
        }

    # Load workbook
    wb = load_workbook(workbook_path)
    ws_txn = wb["Amex Transactions"]

    # Find next empty row
    next_row = _get_next_row(ws_txn)

    # Copy formula patterns from the row above
    formula_templates: dict[int, str] = {}
    if next_row > 5:
        for col in range(4, 10):  # D through I
            val = ws_txn.cell(row=next_row - 1, column=col).value
            if val and isinstance(val, str) and val.startswith("="):
                formula_templates[col] = val

        # Insert transactions
    for tx in transactions:
        ws_txn.cell(row=next_row, column=1).value = statement_name
        ws_txn.cell(row=next_row, column=2).value = datetime.combine(tx.date, datetime.min.time())
        ws_txn.cell(row=next_row, column=3).value = tx.merchant
        ws_txn.cell(row=next_row, column=6).value = tx.foreign_spend
        ws_txn.cell(row=next_row, column=7).value = tx.amount_hkd
        ws_txn.cell(row=next_row, column=8).value = tx.source_page

        # Auto-categorize if known
        category, belongs_to = _lookup_category(wb, tx.merchant)
        if category:
            ws_txn.cell(row=next_row, column=4).value = category
            ws_txn.cell(row=next_row, column=5).value = belongs_to or "Review"
        else:
            # Apply formulas (Category, Belongs To, Review)
            for col, formula in formula_templates.items():
                # Adjust row references
                new_formula = re.sub(r"(\D)(\d+)", lambda m: m.group(1) + str(int(m.group(2)) + 1) if int(m.group(2)) >= (next_row - 1) else m.group(0), formula)
                ws_txn.cell(row=next_row, column=col).value = new_formula

        # Add to Review Items and Rule Updates if new merchant
        _add_to_review_items(wb, tx.merchant, tx.date, tx.amount_hkd, statement_name)
        _add_to_rule_updates(wb, tx.merchant)

        next_row += 1

    # Save
    out_path = output_path or workbook_path
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    return {
        "statement_name": statement_name,
        "entries_written": len(transactions),
        "output_path": str(out_path),
        "first_date": transactions[0].date.isoformat(),
        "last_date": transactions[-1].date.isoformat(),
        "credits": sum(1 for t in transactions if t.is_credit),
        "debits": sum(1 for t in transactions if not t.is_credit),
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Convert AMex statement PDF and append to AMEX.xlsx.")
    parser.add_argument("pdf", type=Path, help="Path to AMex statement PDF")
    parser.add_argument("workbook", type=Path, help="Path to AMEX.xlsx workbook")
    parser.add_argument("--output", type=Path, default=None, help="Output path (default: overwrite workbook)")
    args = parser.parse_args()

    result = convert_amex_statement(
        pdf_path=args.pdf,
        workbook_path=args.workbook,
        output_path=args.output,
    )

    print(f"Statement: {result['statement_name']}")
    print(f"Entries: {result['entries_written']}")
    print(f"Credits: {result.get('credits', 0)}")
    print(f"Debits: {result.get('debits', 0)}")
    print(f"Saved: {result['output_path']}")
