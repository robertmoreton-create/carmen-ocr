#!/usr/bin/env python3
"""Azure Document Intelligence → Excel pipeline for HSBC statements.

Replaces the entire OCR+parse chain.  Calls Azure prebuilt-layout, extracts
transaction tables, parses dates/amounts/currencies, and writes directly into
the existing workbook format.
"""

from __future__ import annotations

import os
import re
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# Azure credentials (same fallback chain as azure_ocr_backend)
# ---------------------------------------------------------------------------

AZURE_ENDPOINT = os.environ.get(
    "AZURE_DOCUMENT_INTELLIGENCE_ENDPOINT",
    "https://documentbank.cognitiveservices.azure.com/",
)
AZURE_KEY = os.environ.get("AZURE_DOCUMENT_INTELLIGENCE_KEY", "")

if not AZURE_KEY:
    _cfg_path = Path(__file__).with_name("azure_config.json")
    if _cfg_path.exists():
        import json
        _cfg = json.loads(_cfg_path.read_text())
        AZURE_ENDPOINT = _cfg.get("endpoint", AZURE_ENDPOINT)
        AZURE_KEY = _cfg.get("key", "")
# Production builds should NOT include test_azure.py —
# the key must come from env var or azure_config.json only.
# The dev fallback below is disabled by default; set CARMEN_DEV=1 to enable.
if not AZURE_KEY and os.environ.get("CARMEN_DEV") == "1":
    try:
        import importlib.util
        _test_azure_path = Path(__file__).parent / "test_azure.py"
        if _test_azure_path.exists():
            _spec = importlib.util.spec_from_file_location("test_azure", _test_azure_path)
            _mod = importlib.util.module_from_spec(_spec)
            _spec.loader.exec_module(_mod)
            AZURE_KEY = getattr(_mod, "KEY", "")
            AZURE_ENDPOINT = getattr(_mod, "ENDPOINT", AZURE_ENDPOINT)
    except Exception:
        pass

# ---------------------------------------------------------------------------
# Types
# ---------------------------------------------------------------------------

@dataclass
class Transaction:
    t_date: date | None
    v_date: date | None
    description: str
    cq: str
    cny: float | None
    gbp: float | None
    jpy: float | None
    hkd: float | None
    usd: float | None


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

CURRENCY_COLUMN = {
    "CNY": 5,
    "GBP": 6,
    "JPY": 7,
    "HKD": 8,
    "USD": 9,
}

TARGET_FONT = Font(name="Times New Roman", size=10)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def _parse_year_from_text(text: str) -> int | None:
    for m in re.finditer(r"\b(20\d{2})\b", text):
        year = int(m.group(1))
        if 2020 <= year <= 2035:
            return year
    return None


def _parse_day_month(value: str, year: int) -> date | None:
    text = _norm(value)
    if not text:
        return None
    # "19 Feb" or "3 Mar"
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3,})$", text, re.IGNORECASE)
    if not m:
        return None
    day = int(m.group(1))
    month_str = m.group(2)[:3].upper()
    month = MONTHS.get(month_str)
    if not month:
        return None
    return date(year, month, day)


def _extract_amount(amount_str: str) -> float | None:
    if not amount_str:
        return None
    cleaned = amount_str.replace(",", "").replace(" ", "").strip()
    m = re.search(r"-?\d+\.?\d*", cleaned)
    if not m:
        return None
    try:
        return float(m.group())
    except ValueError:
        return None


def _humanize_description(value: str) -> str:
    text = _norm(value)
    if not text:
        return text

    # Remove embedded statement-date fragments like (21FEB26)
    text = re.sub(r"\(\d{1,2}[A-Z]{3}\d{2}\)", "", text, flags=re.IGNORECASE)
    text = _norm(text)

    # Simplify noisy provider suffixes
    text = re.sub(r"\*?TRIP\s*HELP\.UBER\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\*?TRIPHELP\.UBER\b", "", text, flags=re.IGNORECASE)
    text = _norm(text)

    # Normalize POS prefixes
    text = re.sub(r"^POSMDC\b", "POS MDC", text, flags=re.IGNORECASE)
    text = re.sub(r"^POS\s+MDC\s*S\b", "POS MDC", text, flags=re.IGNORECASE)

    # "POS MDC ... / MERCHANT" -> "POS MDC - MERCHANT"
    if re.match(r"^POS\s+MDC\b", text, flags=re.IGNORECASE) and " / " in text:
        left, right = text.split(" / ", 1)
        right = _norm(right)
        if right:
            text = f"POS MDC - {right}"
        else:
            text = _norm(left)

    # Reader-friendly POS cleanup
    if text.upper().startswith("POS "):
        text = re.sub(r"^POS\s+MDC\s*(-\s*)?", "POS ", text, flags=re.IGNORECASE)
        merchant = text[4:].strip()
        if "UBER" in merchant.upper():
            text = "POS Uber"
        elif merchant.isupper():
            text = f"POS {merchant.title()}"

    text = re.sub(r"\bCASH\s*R[=\-]?BATE\b", "CASH REBATE", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCASHREBATE\b", "CASH REBATE", text, flags=re.IGNORECASE)
    text = re.sub(r"\bCREDITINTEREST\b", "CREDIT INTEREST", text, flags=re.IGNORECASE)

    # Drop Chinese helper tokens
    tokens = []
    for token in text.split():
        if re.search(r"[\u4e00-\u9fff]", token):
            continue
        tokens.append(token)
    text = _norm(" ".join(tokens))

    return text


# ---------------------------------------------------------------------------
# Azure DI call
# ---------------------------------------------------------------------------

def _azure_extract_tables(pdf_path: Path) -> list[list[list[str]]]:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential

    client = DocumentAnalysisClient(AZURE_ENDPOINT, AzureKeyCredential(AZURE_KEY))

    with pdf_path.open("rb") as f:
        poller = client.begin_analyze_document("prebuilt-layout", f)
        result = poller.result()

    tables: list[list[list[str]]] = []
    for table in result.tables:
        grid: dict[tuple[int, int], str] = {}
        for cell in table.cells:
            grid[(cell.row_index, cell.column_index)] = cell.content.strip()
        rows = []
        for r in range(table.row_count):
            row = [grid.get((r, c), "") for c in range(table.column_count)]
            rows.append(row)
        tables.append(rows)

    return tables


# ---------------------------------------------------------------------------
# Table parsing
# ---------------------------------------------------------------------------

def _is_transaction_table(table: list[list[str]]) -> bool:
    if len(table) < 3:
        return False
    # Scan first 3 rows for a header that has Date + amount columns
    for row_idx in range(min(3, len(table))):
        row = table[row_idx]
        if len(row) < 4:
            continue
        header = " ".join(row).lower()
        has_date = "date" in header or "日期" in header
        has_amount_col = any(k in header for k in ["deposit", "withdrawal", "balance", "存入", "支出", "結餘"])
        if has_date and has_amount_col:
            return True
    return False


def _find_column_indices(header: list[str]) -> dict[str, int | None]:
    lower = [_norm(h).lower() for h in header]
    result: dict[str, int | None] = {
        "date": None,
        "desc": None,
        "deposit": None,
        "withdrawal": None,
        "balance": None,
        "ccy": None,
    }
    for i, h in enumerate(lower):
        if "date" in h or "日期" in h:
            result["date"] = i
        elif "detail" in h or "description" in h or "進支詳情" in h:
            result["desc"] = i
        elif "deposit" in h or "存入" in h:
            result["deposit"] = i
        elif "withdrawal" in h or "支出" in h:
            result["withdrawal"] = i
        elif "balance" in h or "結餘" in h:
            result["balance"] = i
        elif "ccy" in h or "貨幣" in h:
            result["ccy"] = i
    return result


def _parse_tables(tables: list[list[list[str]]], year_hint: int) -> tuple[list[Transaction], dict[str, float], date | None]:
    transactions: list[Transaction] = []
    opening_balances: dict[str, float] = {}
    opening_date: date | None = None

    for table in tables:
        if not _is_transaction_table(table):
            continue

        # Detect header row — sometimes row 0 is a page header, real table header is row 1 or 2
        header_row_idx = 0
        header = table[0]
        header_text = " ".join(header).lower()
        # Try up to first 3 rows to find a real header with Date + amount columns
        for try_idx in range(min(3, len(table))):
            try_header = table[try_idx]
            try_text = " ".join(try_header).lower()
            try_cols = _find_column_indices(try_header)
            if try_cols["date"] is not None and (try_cols["deposit"] is not None or try_cols["withdrawal"] is not None or try_cols["balance"] is not None):
                header_row_idx = try_idx
                header = try_header
                header_text = try_text
                break
        else:
            # Fallback: accept any row with "date" in it
            for try_idx in range(min(3, len(table))):
                try_header = table[try_idx]
                try_text = " ".join(try_header).lower()
                if "date" in try_text or "日期" in try_text:
                    header_row_idx = try_idx
                    header = try_header
                    header_text = try_text
                    break

        cols = _find_column_indices(header)
        if cols["date"] is None and cols["desc"] is None:
            continue

        current_date: date | None = None
        pending_desc = ""

        for row in table[header_row_idx + 1:]:
            if len(row) < 3:
                continue

            def cell(idx: int | None) -> str:
                if idx is None or idx >= len(row):
                    return ""
                return _norm(row[idx])

            date_val = cell(cols["date"])
            desc_val = cell(cols["desc"])
            deposit_val = cell(cols["deposit"])
            withdrawal_val = cell(cols["withdrawal"])
            balance_val = cell(cols["balance"])
            ccy_val = cell(cols["ccy"])

            new_date = _parse_day_month(date_val, year_hint) if date_val else None
            if new_date:
                current_date = new_date

            deposit = _extract_amount(deposit_val)
            withdrawal = _extract_amount(withdrawal_val)
            balance = _extract_amount(balance_val)

            # Skip empty rows
            if not desc_val and deposit is None and withdrawal is None:
                continue

            # Opening balance row
            if desc_val and "b/f balance" in desc_val.lower():
                currency = ccy_val.upper() if ccy_val else "HKD"
                if balance is not None:
                    opening_balances[currency] = balance
                if current_date:
                    opening_date = current_date
                continue

            # Skip totals without amounts
            if desc_val and "total" in desc_val.lower() and deposit is None and withdrawal is None:
                continue

            has_amount = deposit is not None or withdrawal is not None
            is_rebate = "CASH REBATE" in desc_val.upper()
            is_pos_marker = "POS MDC" in desc_val.upper() or "MDC S" in desc_val.upper()
            is_credit_advised = "CREDIT AS ADVISED" in desc_val.upper() or "轉賬收入" in desc_val
            is_debit_advised = "轉賬支出" in desc_val or "EFTPAY" in desc_val.upper()

            # Continuation row (no amount, has description)
            if desc_val and not has_amount and not is_credit_advised and not is_debit_advised:
                pending_desc = desc_val
                continue

            # Merge pending description with current row
            final_desc = desc_val
            if has_amount and pending_desc:
                if is_credit_advised or is_debit_advised:
                    if pending_desc.upper() not in desc_val.upper():
                        final_desc = f"{pending_desc} {desc_val}"
                elif desc_val and not is_pos_marker and not is_rebate:
                    final_desc = f"{pending_desc} {desc_val}"
                pending_desc = ""

            # Clean up: if pending was a POS marker and current row has merchant, combine cleanly
            if pending_desc and not has_amount:
                # Still pending — don't use it yet
                final_desc = desc_val if desc_val else pending_desc
                pending_desc = desc_val if desc_val else pending_desc
            else:
                pending_desc = ""

            # Clean up duplicates and Chinese tokens
            final_desc = re.sub(r"(轉賬收入|轉賬支出)\s+\1", r"\1", final_desc)
            final_desc = re.sub(r"(CREDIT AS ADVISED)\s+\1", r"\1", final_desc)
            final_desc = re.sub(r"(POS MDC \([^)]+\)S)\s+轉賬支出\s+", r"\1 ", final_desc)
            final_desc = re.sub(r"(POS MDC \([^)]+\)S)\s+轉賬收入\s+", r"\1 ", final_desc)
            final_desc = re.sub(r"CREDIT AS ADVISED\s+轉賬收入", "CREDIT AS ADVISED", final_desc)
            final_desc = re.sub(r"CREDIT AS ADVISED\s+轉賬支出", "CREDIT AS ADVISED", final_desc)
            # Clean "CASH REBATE 1450 ..." → "CASH REBATE"
            final_desc = re.sub(r"CASH\s+REBATE\s+\d+", "CASH REBATE", final_desc, flags=re.IGNORECASE)
            # Clean "TO PAYME(HSBC) T250223JY219" → "TO PAYME(HSBC)" (strip transaction codes)
            final_desc = re.sub(r"(TO\s+PAYME\(HSBC\))\s+T\d+[A-Z]+\d+", r"\1", final_desc, flags=re.IGNORECASE)
            final_desc = _norm(final_desc)
            final_desc = _humanize_description(final_desc)

            # Determine currency and amounts
            desc_upper = final_desc.upper()
            currency_match = re.search(r"(USD|GBP|JPY|CNY|KRW)[\s]*([\d,]+\.?\d*)", desc_upper)

            hkd_amount: float | None = None
            usd_amount: float | None = None
            gbp_amount: float | None = None
            jpy_amount: float | None = None
            cny_amount: float | None = None

            if currency_match:
                currency_code = currency_match.group(1)
                foreign_amount = float(currency_match.group(2).replace(",", ""))
                if withdrawal:
                    hkd_amount = -abs(withdrawal)
                elif deposit:
                    hkd_amount = deposit
                if currency_code == "USD":
                    usd_amount = -foreign_amount if withdrawal else foreign_amount
                elif currency_code == "GBP":
                    gbp_amount = -foreign_amount if withdrawal else foreign_amount
                elif currency_code in ("JPY", "KRW"):
                    jpy_amount = -foreign_amount if withdrawal else foreign_amount
                elif currency_code == "CNY":
                    cny_amount = -foreign_amount if withdrawal else foreign_amount
            elif "USD" in desc_upper:
                if withdrawal:
                    hkd_amount = -abs(withdrawal)
                    usd_amount = -abs(withdrawal)
                elif deposit:
                    hkd_amount = deposit
                    usd_amount = deposit
            elif "GBP" in desc_upper:
                if withdrawal:
                    hkd_amount = -abs(withdrawal)
                    gbp_amount = -abs(withdrawal)
                elif deposit:
                    hkd_amount = deposit
                    gbp_amount = deposit
            elif "JPY" in desc_upper or "KRW" in desc_upper:
                if withdrawal:
                    hkd_amount = -abs(withdrawal)
                    jpy_amount = -abs(withdrawal)
                elif deposit:
                    hkd_amount = deposit
                    jpy_amount = deposit
            elif "CNY" in desc_upper:
                if withdrawal:
                    hkd_amount = -abs(withdrawal)
                    cny_amount = -abs(withdrawal)
                elif deposit:
                    hkd_amount = deposit
                    cny_amount = deposit
            else:
                if withdrawal:
                    hkd_amount = -abs(withdrawal)
                elif deposit:
                    hkd_amount = deposit

            # Default currency for the row
            row_currency = ccy_val.upper() if ccy_val else "HKD"
            if row_currency not in CURRENCY_COLUMN:
                row_currency = "HKD"

            # If no foreign amount parsed but ccy column says otherwise, use table amount
            if row_currency == "USD" and usd_amount is None and (deposit or withdrawal):
                usd_amount = -abs(withdrawal) if withdrawal else deposit
            elif row_currency == "GBP" and gbp_amount is None and (deposit or withdrawal):
                gbp_amount = -abs(withdrawal) if withdrawal else deposit
            elif row_currency == "JPY" and jpy_amount is None and (deposit or withdrawal):
                jpy_amount = -abs(withdrawal) if withdrawal else deposit
            elif row_currency == "CNY" and cny_amount is None and (deposit or withdrawal):
                cny_amount = -abs(withdrawal) if withdrawal else deposit

            # Ensure at least one amount is set (default to HKD if nothing else)
            if hkd_amount is None and usd_amount is None and gbp_amount is None and jpy_amount is None and cny_amount is None:
                if deposit is not None:
                    hkd_amount = deposit
                elif withdrawal is not None:
                    hkd_amount = -abs(withdrawal)

            transactions.append(
                Transaction(
                    t_date=current_date,
                    v_date=current_date,
                    description=final_desc,
                    cq="",
                    cny=cny_amount,
                    gbp=gbp_amount,
                    jpy=jpy_amount,
                    hkd=hkd_amount,
                    usd=usd_amount,
                )
            )

            if has_amount:
                pending_desc = ""

    return transactions, opening_balances, opening_date


# ---------------------------------------------------------------------------
# Excel writing (same format as hsbc_statement_to_excel)
# ---------------------------------------------------------------------------

def _clear_sheet_data(ws) -> None:
    for row in range(5, ws.max_row + 1):
        for col in range(1, 10):
            ws.cell(row=row, column=col).value = None


def _ensure_target_sheet(workbook, template_sheet_name: str, target_sheet_name: str):
    template_sheet = workbook[template_sheet_name]
    if target_sheet_name in workbook.sheetnames:
        ws = workbook[target_sheet_name]
    else:
        ws = workbook.copy_worksheet(template_sheet)
        ws.title = target_sheet_name
    return ws


def _write_entries(
    ws,
    transactions: list[Transaction],
    opening_balances: dict[str, float],
    opening_date: date | None,
) -> None:
    opening_style_by_col = {col: copy(ws.cell(row=5, column=col)._style) for col in range(1, 10)}
    txn_style_by_col = {col: copy(ws.cell(row=6, column=col)._style) for col in range(1, 10)}

    _clear_sheet_data(ws)

    date_number_format = ws.cell(row=5, column=2).number_format or "mm/dd/yy;@"
    amount_number_formats = {
        col: ws.cell(row=5, column=col).number_format or "#,##0.00"
        for col in CURRENCY_COLUMN.values()
    }

    ws.cell(row=5, column=3).value = "Balance B/F"
    for col in range(1, 10):
        ws.cell(row=5, column=col)._style = copy(opening_style_by_col[col])
        ws.cell(row=5, column=col).font = copy(TARGET_FONT)
    if opening_date:
        ws.cell(row=5, column=2).value = datetime.combine(opening_date, datetime.min.time())
        ws.cell(row=5, column=2).number_format = date_number_format

    for currency, value in opening_balances.items():
        col = CURRENCY_COLUMN.get(currency)
        if col:
            ws.cell(row=5, column=col).value = value
            ws.cell(row=5, column=col).number_format = amount_number_formats[col]

    row = 6
    previous_date: date | None = None
    for tx in transactions:
        for col in range(1, 10):
            ws.cell(row=row, column=col)._style = copy(txn_style_by_col[col])
            ws.cell(row=row, column=col).font = copy(TARGET_FONT)

        if previous_date != tx.t_date:
            if tx.t_date:
                ws.cell(row=row, column=2).value = datetime.combine(tx.t_date, datetime.min.time())
                ws.cell(row=row, column=2).number_format = date_number_format
            previous_date = tx.t_date
        else:
            ws.cell(row=row, column=2).value = None

        ws.cell(row=row, column=3).value = tx.description

        for currency, amount in [
            ("CNY", tx.cny),
            ("GBP", tx.gbp),
            ("JPY", tx.jpy),
            ("HKD", tx.hkd),
            ("USD", tx.usd),
        ]:
            if amount is not None:
                col = CURRENCY_COLUMN[currency]
                ws.cell(row=row, column=col).value = amount
                ws.cell(row=row, column=col).number_format = amount_number_formats[col]

        row += 1

    # Font normalization pass
    for r in range(1, ws.max_row + 1):
        for c in range(1, 10):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).font = copy(TARGET_FONT)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def convert_statement_azure(
    pdf_path: Path,
    template_path: Path,
    output_path: Path,
    template_sheet_name: str | None = None,
    sheet_name: str = "Auto",
) -> dict:
    """End-to-end: PDF → Azure DI → Excel."""

    tables = _azure_extract_tables(pdf_path)

    # Detect year from filename or table content
    year_hint = date.today().year
    basename = pdf_path.name
    m = re.search(r"20(\d{2})", basename)
    if m:
        year_hint = 2000 + int(m.group(1))
    else:
        # Try to find year in first table
        for table in tables:
            for row in table[:3]:
                for cell in row:
                    y = _parse_year_from_text(cell)
                    if y:
                        year_hint = y
                        break
                if year_hint != date.today().year:
                    break
            if year_hint != date.today().year:
                break

    transactions, opening_balances, opening_date = _parse_tables(tables, year_hint)

    workbook = load_workbook(template_path)
    selected_template = template_sheet_name or workbook.sheetnames[0]
    ws = _ensure_target_sheet(workbook, selected_template, sheet_name)
    _write_entries(ws, transactions, opening_balances, opening_date)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "statement_year": year_hint,
        "opening_balances": opening_balances,
        "entries_written": len(transactions),
        "output_path": str(output_path),
        "sheet_name": sheet_name,
        "opening_date": opening_date.isoformat() if opening_date else None,
        "parser_engine": "azure-di",
        "has_any_entries": bool(transactions),
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Convert scanned HSBC statement PDF to Excel via Azure DI.")
    parser.add_argument("pdf", type=Path, help="Path to scanned statement PDF")
    parser.add_argument("template", type=Path, help="Path to existing Excel workbook template")
    parser.add_argument("output", type=Path, help="Path for output workbook")
    parser.add_argument("--template-sheet", default=None, help="Sheet to copy formatting from")
    parser.add_argument("--sheet-name", default="Auto", help="Output sheet name")
    args = parser.parse_args()

    result = convert_statement_azure(
        pdf_path=args.pdf,
        template_path=args.template,
        output_path=args.output,
        template_sheet_name=args.template_sheet,
        sheet_name=args.sheet_name,
    )

    print(f"Statement year: {result['statement_year']}")
    print(f"Opening balances: {result['opening_balances']}")
    print(f"Entries written: {result['entries_written']}")
    print(f"Saved: {result['output_path']}")
